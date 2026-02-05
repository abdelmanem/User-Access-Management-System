import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib.auth.forms import SetPasswordForm
from django.db import transaction
from django.db.models import Q, F, Value
from django.db.models.functions import Concat
from django.utils import timezone
from .models import CustomUser, UserDeactivationAudit, UserArchive, LDAPConfiguration
from .forms import (
    UserCreateForm,
    UserUpdateForm,
    UserPermissionForm,
    UserPhotoForm,
    LDAPConfigurationForm,
    LDAPTestConnectionForm,
    LDAPTestLoginForm,
    LDAPSyncForm,
)
from departments.models import Department
from hardware.models import HardwareAsset
from urllib.parse import urlencode
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from access_management.models import UserSystemAccess
from PIL import Image, UnidentifiedImageError
import io

@login_required
@user_passes_test(lambda u: u.is_staff)
def user_list(request):
    from django.core.paginator import Paginator

    users_qs = CustomUser.objects.select_related('department').all().annotate(
        sort_full_name=Concat(
            F('first_name'),
            Value(' '),
            F('last_name'),
        )
    )

    # Filters
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()  # 'active', 'inactive', or ''
    dept_id = request.GET.get('department', '').strip()
    follow_up = request.GET.get('follow_up', '').strip()
    page_size_param = request.GET.get('page_size', '').strip().lower()
    metrics_filter = request.GET.get('metrics', '').strip().lower()
    allowed_page_sizes = [25, 50, 100]
    paginate = True
    if page_size_param == 'all':
        paginate = False
        page_size = None
        page_size_display = 'all'
    else:
        try:
            page_size = int(page_size_param) if page_size_param else 25
        except ValueError:
            page_size = 25
        if page_size not in allowed_page_sizes:
            page_size = 25
        page_size_display = page_size

    if q:
        users_qs = users_qs.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q) |
            Q(employee_id__icontains=q)
        )
    if status == 'active':
        users_qs = users_qs.filter(is_active=True)
    elif status == 'inactive':
        users_qs = users_qs.filter(is_active=False)
    if dept_id:
        if dept_id == 'none':
            users_qs = users_qs.filter(department__isnull=True)
        else:
            users_qs = users_qs.filter(department_id=dept_id)
    if follow_up == 'flagged':
        users_qs = users_qs.filter(flag_for_follow_up=True)
    elif follow_up == 'unflagged':
        users_qs = users_qs.filter(flag_for_follow_up=False)
    if metrics_filter == 'included':
        users_qs = users_qs.filter(exclude_from_metrics=False)
    elif metrics_filter == 'excluded':
        users_qs = users_qs.filter(exclude_from_metrics=True)

    sort_key = (request.GET.get('sort') or 'full_name').strip()
    sort_dir = (request.GET.get('dir') or 'asc').strip().lower()

    sort_map = {
        'id': 'id',
        'username': 'username',
        'full_name': 'sort_full_name',
        'name': 'sort_full_name',
        'email': 'email',
        'department': 'department__name',
        'position': 'position',
        'description': 'description',
        'follow_up': 'flag_for_follow_up',
        'metrics': 'exclude_from_metrics',
        'notes': 'notes',
        'status': 'is_active',
    }

    order_field = sort_map.get(sort_key, 'full_name')
    if sort_dir == 'desc':
        order_field = f'-{order_field}'

    users_qs = users_qs.order_by(order_field, 'id')

    total_count = users_qs.count()
    reportable_kpi_qs = users_qs.filter(exclude_from_metrics=False)
    excluded_kpi_qs = users_qs.filter(exclude_from_metrics=True)

    included_count = reportable_kpi_qs.count()
    excluded_count = excluded_kpi_qs.count()
    active_reportable_count = reportable_kpi_qs.filter(is_active=True).count()
    inactive_reportable_count = reportable_kpi_qs.filter(is_active=False).count()
    no_department_count = reportable_kpi_qs.filter(department__isnull=True).count()
    follow_up_count = reportable_kpi_qs.filter(flag_for_follow_up=True).count()

    active_count = active_reportable_count
    inactive_count = inactive_reportable_count

    # Helper function to check if current filters match KPI filter
    def matches_kpi_filter(kpi_filter):
        # Check that all KPI filter params match current filters
        for key, value in kpi_filter.items():
            if key == 'metrics':
                if metrics_filter != value:
                    return False
            elif key == 'status':
                if status != value:
                    return False
            elif key == 'department':
                if dept_id != value:
                    return False
            elif key == 'follow_up':
                if follow_up != value:
                    return False
        # Check that no other filters are active (except search 'q')
        if metrics_filter and 'metrics' not in kpi_filter:
            return False
        if status and 'status' not in kpi_filter:
            return False
        if dept_id and 'department' not in kpi_filter:
            return False
        if follow_up and 'follow_up' not in kpi_filter:
            return False
        return True

    user_kpis = [
        {
            'label': 'Reportable Users',
            'value': included_count,
            'badge': 'primary',
            'description': 'Included in KPI metrics',
            'filter_params': json.dumps({'metrics': 'included'}),
            'is_active': matches_kpi_filter({'metrics': 'included'}),
        },
        {
            'label': 'Active Reportable Users',
            'value': active_reportable_count,
            'badge': 'success',
            'description': 'Reportable accounts currently active',
            'filter_params': json.dumps({'metrics': 'included', 'status': 'active'}),
            'is_active': matches_kpi_filter({'metrics': 'included', 'status': 'active'}),
        },
        {
            'label': 'Excluded from Metrics',
            'value': excluded_count,
            'badge': 'secondary',
            'description': 'Users intentionally excluded from reporting',
            'filter_params': json.dumps({'metrics': 'excluded'}),
            'is_active': matches_kpi_filter({'metrics': 'excluded'}),
        },
        {
            'label': 'Active',
            'value': active_count,
            'badge': 'info',
            'description': 'All users marked active in this view',
            'filter_params': json.dumps({'status': 'active'}),
            'is_active': matches_kpi_filter({'status': 'active'}),
        },
        {
            'label': 'Inactive',
            'value': inactive_count,
            'badge': 'warning',
            'description': 'All users marked inactive in this view',
            'filter_params': json.dumps({'status': 'inactive'}),
            'is_active': matches_kpi_filter({'status': 'inactive'}),
        },
        {
            'label': 'No Department',
            'value': no_department_count,
            'badge': 'danger',
            'description': 'Users missing department assignment',
            'filter_params': json.dumps({'department': 'none'}),
            'is_active': matches_kpi_filter({'department': 'none'}),
        },
        {
            'label': 'Need Follow-up',
            'value': follow_up_count,
            'badge': 'dark',
            'description': 'Flagged for manual review',
            'filter_params': json.dumps({'follow_up': 'flagged'}),
            'is_active': matches_kpi_filter({'follow_up': 'flagged'}),
        },
    ]

    if paginate:
        paginator = Paginator(users_qs, page_size)  # type: ignore[arg-type]
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        users_page = page_obj.object_list
    else:
        page_obj = None
        users_page = list(users_qs)

    departments = Department.objects.all().order_by('name')

    query_params = request.GET.copy()
    for key in ['sort', 'dir', 'page']:
        if key in query_params:
            del query_params[key]
    base_query = query_params.urlencode()

    return render(request, 'accounts/user_list.html', {
        'page_obj': page_obj,
        'users': users_page,
        'q': q,
        'status': status,
        'department_selected': dept_id,
        'follow_up': follow_up,
        'metrics': metrics_filter,
        'departments': departments,
        'page_size': page_size_display,
        'allowed_page_sizes': allowed_page_sizes,
        'total_count': total_count,
        'current_sort': sort_key,
        'current_dir': 'desc' if sort_dir == 'desc' else 'asc',
        'base_query': base_query,
        'active_count': active_count,
        'inactive_count': inactive_count,
        'no_department_count': no_department_count,
        'follow_up_count': follow_up_count,
        'excluded_count': excluded_count,
        'included_count': included_count,
        'active_reportable_count': active_reportable_count,
        'user_kpis': user_kpis,
    })


def _build_filtered_users_queryset(request):
    users_qs = CustomUser.objects.select_related('department').all().annotate(
        sort_full_name=Concat(
            F('first_name'),
            Value(' '),
            F('last_name'),
        )
    )
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    dept_id = request.GET.get('department', '').strip()
    follow_up = request.GET.get('follow_up', '').strip()
    metrics_filter = request.GET.get('metrics', '').strip().lower()
    if q:
        users_qs = users_qs.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q) |
            Q(employee_id__icontains=q)
        )
    if status == 'active':
        users_qs = users_qs.filter(is_active=True)
    elif status == 'inactive':
        users_qs = users_qs.filter(is_active=False)
    if dept_id:
        if dept_id == 'none':
            users_qs = users_qs.filter(department__isnull=True)
        else:
            users_qs = users_qs.filter(department_id=dept_id)
    if follow_up == 'flagged':
        users_qs = users_qs.filter(flag_for_follow_up=True)
    elif follow_up == 'unflagged':
        users_qs = users_qs.filter(flag_for_follow_up=False)
    if metrics_filter == 'included':
        users_qs = users_qs.filter(exclude_from_metrics=False)
    elif metrics_filter == 'excluded':
        users_qs = users_qs.filter(exclude_from_metrics=True)

    sort_key = (request.GET.get('sort') or 'full_name').strip()
    sort_dir = (request.GET.get('dir') or 'asc').strip().lower()
    sort_map = {
        'id': 'id',
        'username': 'username',
        'full_name': 'sort_full_name',
        'name': 'sort_full_name',
        'email': 'email',
        'department': 'department__name',
        'position': 'position',
        'description': 'description',
        'follow_up': 'flag_for_follow_up',
        'metrics': 'exclude_from_metrics',
        'notes': 'notes',
        'status': 'is_active',
    }
    order_field = sort_map.get(sort_key, 'sort_full_name')
    if sort_dir == 'desc':
        order_field = f'-{order_field}'
    return users_qs.order_by(order_field, 'id')


def _filtered_deactivation_audits(query: str = '', start_date=None, end_date=None):
    audits = UserDeactivationAudit.objects.select_related('user', 'admin').order_by('-deactivated_at')
    if query:
        audits = audits.filter(
            Q(user_username__icontains=query)
            | Q(user_full_name__icontains=query)
            | Q(user_employee_id__icontains=query)
            | Q(admin__username__icontains=query)
            | Q(admin__first_name__icontains=query)
            | Q(admin__last_name__icontains=query)
        )
    if start_date:
        audits = audits.filter(deactivated_at__date__gte=start_date)
    if end_date:
        audits = audits.filter(deactivated_at__date__lte=end_date)
    return audits


def _filtered_user_archives(query: str = '', start_date=None, end_date=None):
    archives = UserArchive.objects.select_related('archived_by').order_by('-archived_at')
    if query:
        archives = archives.filter(
            Q(username__icontains=query)
            | Q(full_name__icontains=query)
            | Q(employee_id__icontains=query)
            | Q(department_name__icontains=query)
        )
    if start_date:
        archives = archives.filter(archived_at__date__gte=start_date)
    if end_date:
        archives = archives.filter(archived_at__date__lte=end_date)
    return archives


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_export_excel(request):
    qs = _build_filtered_users_queryset(request)
    headers = [
        'ID',
        'Username',
        'Full Name',
        'Email',
        'Department',
        'Position',
        'Status',
        'Follow-up Flag',
        'Excluded from Metrics',
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Users'
    ws.append(headers)
    for u in qs:
        ws.append([
            u.id,
            u.username,
            f"{u.first_name} {u.last_name}".strip(),
            u.email,
            u.department.name if u.department else '',
            u.position or '',
            'Active' if u.is_active else 'Inactive',
            'Yes' if u.flag_for_follow_up else 'No',
            'Yes' if u.exclude_from_metrics else 'No',
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_export_pdf(request):
    qs = _build_filtered_users_queryset(request)
    data = [[
        'ID',
        'Username',
        'Full Name',
        'Email',
        'Department',
        'Position',
        'Status',
        'Follow-up Flag',
        'Excluded from Metrics',
    ]]
    for u in qs:
        data.append([
            str(u.id),
            u.username,
            f"{u.first_name} {u.last_name}".strip(),
            u.email,
            u.department.name if u.department else '',
            u.position or '',
            'Active' if u.is_active else 'Inactive',
            'Yes' if u.flag_for_follow_up else 'No',
            'Yes' if u.exclude_from_metrics else 'No',
        ])
    buffer_response = HttpResponse(content_type='application/pdf')
    buffer_response['Content-Disposition'] = 'attachment; filename="users.pdf"'
    page_size = landscape(A4)
    c = canvas.Canvas(buffer_response, pagesize=page_size)
    width, height = page_size
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))
    # Calculate table size and paginate if needed
    available_width = width - 40
    available_height = height - 60
    table.wrapOn(c, available_width, available_height)
    # Simple pagination if table is long
    y = height - 30
    rows_per_page = 25
    header = data[0]
    chunk = []
    for i, row in enumerate(data[1:], start=1):
        chunk.append(row)
        if len(chunk) == rows_per_page or i == len(data) - 1:
            t = Table([header] + chunk, repeatRows=1)
            t.setStyle(table._argW[1]) if hasattr(table, '_argW') else None
            t.setStyle(table._cellStyles) if hasattr(table, '_cellStyles') else None
            t.setStyle(table._tblStyle)
            t.wrapOn(c, available_width, available_height)
            t.drawOn(c, 20, 20)
            c.showPage()
            chunk = []
    c.save()
    return buffer_response

@login_required
@user_passes_test(lambda u: u.is_staff)
def user_detail(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    accesses = UserSystemAccess.objects.filter(user=user).select_related('system').order_by('-created_at')
    return render(request, 'accounts/user_detail.html', {'user': user, 'accesses': accesses})


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_photo_update(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method != 'POST':
        messages.error(request, 'Invalid request method for updating profile photos.')
        return redirect('accounts:user_detail', pk=pk)

    if 'profile_photo' not in request.FILES:
        messages.error(request, 'Please choose an image to upload.')
        return redirect('accounts:user_detail', pk=pk)

    # Validate and save uploaded file directly from request.FILES to avoid
    # validators or other code that may close the uploaded file object.
    uploaded = request.FILES.get('profile_photo')
    if not uploaded:
        messages.error(request, 'Please choose an image to upload.')
        return redirect('accounts:user_detail', pk=pk)

    # Size check
    max_mb = 2
    if getattr(uploaded, 'size', 0) > max_mb * 1024 * 1024:
        messages.error(request, f'Profile photo must be <= {max_mb}MB.')
        return redirect('accounts:user_detail', pk=pk)

    # Validate image via PIL from an in-memory copy
    try:
        data = uploaded.read()
        buf = io.BytesIO(data)
        img = Image.open(buf)
        img.verify()
        fmt = getattr(img, 'format', None)
        if fmt not in ['JPEG', 'PNG']:
            messages.error(request, 'Profile photo must be a JPG or PNG image.')
            return redirect('accounts:user_detail', pk=pk)
    except (UnidentifiedImageError, OSError):
        messages.error(request, 'Invalid image file. Please upload a valid JPG or PNG image.')
        return redirect('accounts:user_detail', pk=pk)
    finally:
        try:
            buf.close()
        except Exception:
            pass

    # Save via ContentFile (avoid using the UploadedFile object directly)
    if user.profile_photo:
        user.profile_photo.delete(save=False)
    try:
        file_name = uploaded.name
        user.profile_photo.save(file_name, ContentFile(data), save=False)
        user.updated_by = request.user
        user.save(update_fields=['profile_photo', 'updated_by'])
        messages.success(request, 'Profile photo updated successfully.')
    except Exception:
        messages.error(request, 'Unable to save profile photo.')
    return redirect('accounts:user_detail', pk=pk)


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_photo_delete(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method != 'POST':
        messages.error(request, 'Invalid request method for deleting profile photos.')
        return redirect('accounts:user_detail', pk=pk)

    if user.profile_photo:
        user.profile_photo.delete(save=False)
        user.profile_photo = None
        user.updated_by = request.user
        user.save(update_fields=['profile_photo', 'updated_by'])
        messages.success(request, 'Profile photo removed.')
    else:
        messages.info(request, 'This user does not have a profile photo to remove.')
    return redirect('accounts:user_detail', pk=pk)


@login_required
@permission_required('accounts.add_customuser', raise_exception=True)
def user_create(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.created_by = request.user
            user.updated_by = request.user
            user.save()
            form.save_m2m()
            messages.success(request, 'User created successfully.')
            return redirect('accounts:user_detail', pk=user.pk)
        messages.error(request, 'Please correct the errors below.')
    else:
        form = UserCreateForm()
    return render(request, 'accounts/user_form.html', {'form': form})

@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_update(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            updated_user = form.save(commit=False)
            updated_user.updated_by = request.user
            updated_user.save()
            form.save_m2m()
            messages.success(request, 'User updated successfully.')
            return redirect('accounts:user_detail', pk=user.pk)
        messages.error(request, 'Please correct the errors below.')
    else:
        form = UserUpdateForm(instance=user)
    return render(request, 'accounts/user_form.html', {'form': form, 'user': user})


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_reset_password(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        form = SetPasswordForm(user=user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Password reset successfully.')
            return redirect('accounts:user_detail', pk=user.pk)
        messages.error(request, 'Please correct the errors below.')
    else:
        form = SetPasswordForm(user=user)
    return render(request, 'accounts/user_reset_password.html', {'form': form, 'user': user})

@login_required
@permission_required('accounts.delete_customuser', raise_exception=True)
def user_delete(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    system_assignments = list(UserSystemAccess.objects.filter(user=user).select_related('system'))
    hardware_primary = list(HardwareAsset.objects.filter(primary_user=user))
    hardware_shared = list(HardwareAsset.objects.filter(assigned_users=user))

    if request.method == 'POST':
        archive_payload = {
            'user': {
                'username': user.username,
                'full_name': user.get_full_name(),
                'employee_id': user.employee_id,
                'email': user.email,
                'department': user.department.name if user.department else None,
                'position': user.position,
                'employment_status': user.employment_status,
                'employment_type': user.employment_type,
                'join_date': user.join_date.isoformat() if user.join_date else None,
                'created_at': user.created_at.isoformat() if hasattr(user, 'created_at') and user.created_at else None,
                'updated_at': user.updated_at.isoformat() if hasattr(user, 'updated_at') and user.updated_at else None,
            },
            'system_assignments': [
                {
                    'system_id': assignment.system_id,
                    'system_name': assignment.system.name if assignment.system else None,
                    'system_code': assignment.system.code if assignment.system else None,
                    'status': assignment.status,
                    'access_type': assignment.access_type,
                    'request_type': assignment.request_type,
                    'priority': assignment.priority,
                    'start_date': assignment.access_start_date.isoformat() if assignment.access_start_date else None,
                    'end_date': assignment.access_end_date.isoformat() if assignment.access_end_date else None,
                }
                for assignment in system_assignments
            ],
            'hardware': {
                'primary': [
                    {
                        'asset_id': asset.id,
                        'asset_name': asset.name,
                        'asset_tag': asset.asset_tag,
                        'status': asset.status,
                        'department': asset.department.name if asset.department else None,
                    }
                    for asset in hardware_primary
                ],
                'shared': [
                    {
                        'asset_id': asset.id,
                        'asset_name': asset.name,
                        'asset_tag': asset.asset_tag,
                        'status': asset.status,
                        'department': asset.department.name if asset.department else None,
                    }
                    for asset in hardware_shared
                ],
            },
        }

        UserArchive.objects.create(
            source_user_id=user.id,
            username=user.username,
            full_name=user.get_full_name(),
            employee_id=user.employee_id or '',
            email=user.email or '',
            department_name=user.department.name if user.department else '',
            archived_by=request.user if request.user.is_authenticated else None,
            payload=archive_payload,
        )

        user.delete()
        messages.success(request, 'User deleted successfully. Archived snapshot is available for reference.')
        return redirect('accounts:user_list')

    return render(
        request,
        'accounts/user_confirm_delete.html',
        {
            'user': user,
            'system_assignments': system_assignments,
            'hardware_primary': hardware_primary,
            'hardware_shared': hardware_shared,
        },
    )

@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_toggle_active(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('accounts:user_list')

    if request.method != 'POST':
        return redirect(next_url)

    target_status = not user.is_active
    audit_record_created = False
    system_snapshot = []
    hardware_snapshot = []
    hardware_status_action = 'no_change'

    if user.is_active and not target_status:
        active_assignments = list(UserSystemAccess.objects.filter(
            user=user,
            status__in=['Active', 'Approved', 'Pending']
        ).select_related('system'))

        hardware_primary = list(HardwareAsset.objects.filter(primary_user=user))
        hardware_shared = list(HardwareAsset.objects.filter(assigned_users=user).exclude(primary_user=user))

        hardware_assets_map = {}
        for asset in hardware_primary:
            data = hardware_assets_map.setdefault(
                asset.id,
                {
                    'asset': asset,
                    'roles': set(),
                    'status_before': asset.status,
                    'status_after': asset.status,
                    'was_primary': False,
                    'was_shared': False,
                },
            )
            data['roles'].add('Primary Owner')
            data['was_primary'] = True
        for asset in hardware_shared:
            data = hardware_assets_map.setdefault(
                asset.id,
                {
                    'asset': asset,
                    'roles': set(),
                    'status_before': asset.status,
                    'status_after': asset.status,
                    'was_primary': False,
                    'was_shared': False,
                },
            )
            data['roles'].add('Shared User')
            data['was_shared'] = True
        hardware_assets = []
        for entry in hardware_assets_map.values():
            entry['roles'] = sorted(entry['roles'])
            hardware_assets.append(entry)

        needs_confirmation = bool(active_assignments or hardware_assets)
        confirmed = request.POST.get('confirm_deactivate') == '1'

        if needs_confirmation and not confirmed:
            context = {
                'user': user,
                'next_url': next_url,
                'system_assignments': active_assignments,
                'hardware_assets': hardware_assets,
            }
            return render(request, 'accounts/user_deactivate_confirm.html', context)

        if needs_confirmation and confirmed:
            errors = []
            if active_assignments and request.POST.get('confirm_system', '') != '1':
                errors.append('Please confirm system assignments should be suspended.')
            if hardware_assets and request.POST.get('confirm_hardware', '') != '1':
                errors.append('Please confirm any assigned hardware has been collected.')

            if errors:
                for error in errors:
                    messages.error(request, error)
                context = {
                    'user': user,
                    'next_url': next_url,
                    'system_assignments': active_assignments,
                    'hardware_assets': hardware_assets,
                }
                return render(request, 'accounts/user_deactivate_confirm.html', context)

            system_snapshot = [
                {
                    'system_id': assignment.system_id,
                    'system_name': assignment.system.name if assignment.system else None,
                    'access_type': assignment.access_type,
                    'status_before': assignment.status,
                }
                for assignment in active_assignments
            ]

            hardware_status_action = request.POST.get('hardware_status_action', 'no_change') if hardware_assets else 'no_change'

            with transaction.atomic():
                if active_assignments:
                    now = timezone.now()
                    UserSystemAccess.objects.filter(
                        user=user,
                        status__in=['Active', 'Approved']
                    ).update(status='Suspended', access_end_date=now)
                    UserSystemAccess.objects.filter(
                        user=user,
                        status='Pending'
                    ).update(status='Cancelled')

                    for snapshot in system_snapshot:
                        before = snapshot['status_before']
                        if before in ['Active', 'Approved']:
                            snapshot['status_after'] = 'Suspended'
                        elif before == 'Pending':
                            snapshot['status_after'] = 'Cancelled'
                        else:
                            snapshot['status_after'] = before

                if hardware_assets:
                    for entry in hardware_assets:
                        asset = entry['asset']
                        fields_to_update = set()

                        if entry.get('was_primary'):
                            asset.primary_user = None
                            fields_to_update.add('primary_user')

                        if asset.assigned_users.filter(id=user.id).exists():
                            asset.assigned_users.remove(user)

                        if hardware_status_action == 'in_storage' and asset.status != 'In Storage':
                            asset.status = 'In Storage'
                            fields_to_update.add('status')
                        elif hardware_status_action == 'retired' and asset.status != 'Retired':
                            asset.status = 'Retired'
                            fields_to_update.add('status')

                        if fields_to_update:
                            if hasattr(asset, 'updated_by'):
                                asset.updated_by = request.user
                                fields_to_update.add('updated_by')
                            asset.save(update_fields=list(fields_to_update))
                        entry['status_after'] = asset.status

                    hardware_snapshot = [
                        {
                            'asset_id': entry['asset'].id,
                            'asset_name': entry['asset'].name,
                            'asset_tag': entry['asset'].asset_tag,
                            'roles': entry['roles'],
                            'status_before': entry.get('status_before'),
                            'status_after': entry.get('status_after', entry.get('status_before')),
                            'was_primary': entry.get('was_primary', False),
                            'was_shared': entry.get('was_shared', False),
                        }
                        for entry in hardware_assets
                    ]

                user.is_active = False
                if hasattr(user, 'updated_by'):
                    user.updated_by = request.user
                    user.save(update_fields=['is_active', 'updated_by'])
                else:
                    user.save(update_fields=['is_active'])

                UserDeactivationAudit.objects.create(
                    user=user,
                    user_username=user.username,
                    user_full_name=user.get_full_name(),
                    user_employee_id=user.employee_id or '',
                    admin=request.user if request.user.is_authenticated else None,
                    system_confirmed=bool(active_assignments),
                    hardware_confirmed=bool(hardware_assets),
                    hardware_status_action=hardware_status_action if hardware_assets else 'no_change',
                    system_assignments=system_snapshot,
                    hardware_assignments=hardware_snapshot,
                )
                audit_record_created = True

                messages.success(request, 'User deactivated. Related assignments have been updated.')
                return redirect(next_url)

    user.is_active = target_status
    if hasattr(user, 'updated_by'):
        user.updated_by = request.user
        user.save(update_fields=['is_active', 'updated_by'])
    else:
        user.save(update_fields=['is_active'])

    if not target_status and not audit_record_created:
        UserDeactivationAudit.objects.create(
            user=user,
            user_username=user.username,
            user_full_name=user.get_full_name(),
            user_employee_id=user.employee_id or '',
            admin=request.user if request.user.is_authenticated else None,
            system_confirmed=False,
            hardware_confirmed=False,
            hardware_status_action='not_applicable',
            system_assignments=[],
            hardware_assignments=[],
        )

    messages.success(request, f"User {'activated' if user.is_active else 'deactivated'} successfully.")
    return redirect(next_url)


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_toggle_follow_up(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('accounts:user_list')

    if request.method != 'POST':
        return redirect(next_url)

    user.flag_for_follow_up = not user.flag_for_follow_up
    if hasattr(user, 'updated_by'):
        user.updated_by = request.user
        user.save(update_fields=['flag_for_follow_up', 'updated_by'])
    else:
        user.save(update_fields=['flag_for_follow_up'])

    if user.flag_for_follow_up:
        messages.success(request, f"{user.get_full_name() or user.username} marked for follow-up.")
    else:
        messages.success(request, f"Follow-up flag cleared for {user.get_full_name() or user.username}.")
    return redirect(next_url)


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_toggle_metrics(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('accounts:user_list')

    if request.method != 'POST':
        return redirect(next_url)

    user.exclude_from_metrics = not user.exclude_from_metrics
    update_fields = ['exclude_from_metrics']
    if hasattr(user, 'updated_by'):
        user.updated_by = request.user
        update_fields.append('updated_by')
    user.save(update_fields=update_fields)

    if user.exclude_from_metrics:
        messages.success(request, f"{user.get_full_name() or user.username} will be excluded from dashboards and totals.")
    else:
        messages.success(request, f"{user.get_full_name() or user.username} is now counted in dashboards and totals.")
    return redirect(next_url)

@login_required
@user_passes_test(lambda u: u.is_staff)
def user_bulk_action(request):
    if request.method != 'POST':
        return redirect('accounts:user_list')

    action = request.POST.get('action', '').strip()
    selected_ids = request.POST.getlist('selected')

    # Preserve filters after redirect
    redirect_params = {
        'q': request.POST.get('q', ''),
        'status': request.POST.get('status', ''),
        'department': request.POST.get('department', ''),
        'follow_up': request.POST.get('follow_up', ''),
        'metrics': request.POST.get('metrics', ''),
        'page_size': request.POST.get('page_size', ''),
        'sort': request.POST.get('sort', ''),
        'dir': request.POST.get('dir', ''),
    }
    base_url = reverse('accounts:user_list')
    query_string = urlencode({k: v for k, v in redirect_params.items() if v})
    redirect_url = f"{base_url}?{query_string}" if query_string else base_url

    if not selected_ids:
        messages.warning(request, 'No users selected.')
        return redirect(redirect_url)

    queryset = CustomUser.objects.filter(id__in=selected_ids)

    if action == 'activate':
        if not request.user.has_perm('accounts.change_customuser'):
            messages.error(request, 'You do not have permission to activate users.')
            return redirect(redirect_url)
        updated = queryset.update(is_active=True)
        messages.success(request, f'Activated {updated} user(s).')
    elif action == 'deactivate':
        if not request.user.has_perm('accounts.change_customuser'):
            messages.error(request, 'You do not have permission to deactivate users.')
            return redirect(redirect_url)
        dependency_users = queryset.filter(
            Q(system_accesses__status__in=['Active', 'Approved', 'Pending']) |
            Q(primary_hardware_assets__isnull=False) |
            Q(hardware_assets__isnull=False)
        ).distinct()
        if dependency_users.exists():
            sample = ', '.join(dependency_users.values_list('username', flat=True)[:5])
            if dependency_users.count() > 5:
                sample += ', …'
            messages.error(
                request,
                f"Bulk deactivation halted. The following users still have system or hardware assignments: {sample}. "
                "Please review and deactivate these accounts individually so dependencies can be addressed."
            )
            return redirect(redirect_url)
        updated = queryset.update(is_active=False)
        messages.success(request, f'Deactivated {updated} user(s).')
    elif action == 'delete':
        if not request.user.has_perm('accounts.delete_customuser'):
            messages.error(request, 'You do not have permission to delete users.')
            return redirect(redirect_url)
        count = queryset.count()
        queryset.delete()
        messages.success(request, f'Deleted {count} user(s).')
    elif action == 'assign_department':
        if not request.user.has_perm('accounts.change_customuser'):
            messages.error(request, 'You do not have permission to assign departments.')
            return redirect(redirect_url)
        department_id = request.POST.get('bulk_department_id', '').strip()
        if department_id:
            department = Department.objects.filter(id=department_id).first()
            if department is None:
                messages.error(request, 'Selected department was not found.')
                return redirect(redirect_url)
            updated = queryset.update(department=department, updated_by=request.user)
            messages.success(request, f'Assigned {updated} user(s) to {department.name}.')
        else:
            # Clear department assignment
            updated = queryset.update(department=None, updated_by=request.user)
            messages.success(request, f'Removed department assignment for {updated} user(s).')
    elif action == 'clear_email':
        if not request.user.has_perm('accounts.change_customuser'):
            messages.error(request, 'You do not have permission to modify user emails.')
            return redirect(redirect_url)
        cleared = queryset.update(email='', updated_by=request.user)
        messages.success(request, f'Cleared email addresses for {cleared} user(s).')
    elif action == 'flag_follow_up':
        if not request.user.has_perm('accounts.change_customuser'):
            messages.error(request, 'You do not have permission to flag users.')
            return redirect(redirect_url)
        updated = queryset.update(flag_for_follow_up=True, updated_by=request.user)
        messages.success(request, f'Marked {updated} user(s) for follow-up.')
    elif action == 'clear_follow_up':
        if not request.user.has_perm('accounts.change_customuser'):
            messages.error(request, 'You do not have permission to flag users.')
            return redirect(redirect_url)
        updated = queryset.update(flag_for_follow_up=False, updated_by=request.user)
        messages.success(request, f'Cleared the follow-up flag for {updated} user(s).')
    else:
        messages.warning(request, 'Please choose a valid bulk action.')

    return redirect(redirect_url)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_manage_permissions(request, pk):
    """Allow superusers to manage a user's group memberships and direct permissions."""
    target_user = get_object_or_404(CustomUser, pk=pk)

    if request.method == 'POST':
        form = UserPermissionForm(request.POST, user_instance=target_user)
        if form.is_valid():
            is_staff = form.cleaned_data['is_staff']
            is_superuser_flag = form.cleaned_data['is_superuser']
            groups = form.cleaned_data['groups']
            permissions = form.cleaned_data['permissions']

            if target_user.is_superuser and not is_superuser_flag:
                other_superusers = CustomUser.objects.filter(is_superuser=True).exclude(pk=target_user.pk).exists()
                if not other_superusers:
                    messages.error(request, 'At least one superuser account must remain. Promote another superuser before demoting this user.')
                    return redirect('accounts:user_manage_permissions', pk=target_user.pk)

            target_user.groups.set(groups)
            target_user.user_permissions.set(permissions)
            target_user.is_staff = is_staff
            target_user.is_superuser = is_superuser_flag
            target_user.updated_by = request.user
            target_user.save(update_fields=['is_staff', 'is_superuser', 'updated_by', 'updated_at'])

            messages.success(request, 'Permissions updated successfully.')
            return redirect('accounts:user_detail', pk=target_user.pk)

        messages.error(request, 'Please correct the errors below.')
    else:
        form = UserPermissionForm(user_instance=target_user)

    direct_permissions = target_user.user_permissions.select_related('content_type').order_by(
        'content_type__app_label', 'codename'
    )

    return render(
        request,
        'accounts/user_permissions.html',
        {
            'form': form,
            'user': target_user,
            'direct_permissions': direct_permissions,
        },
    )


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def user_assign_department(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method != 'POST':
        return redirect(request.META.get('HTTP_REFERER', reverse('accounts:user_list')))

    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('accounts:user_list')
    department_id = (request.POST.get('department_id') or '').strip()

    if department_id:
        department = Department.objects.filter(id=department_id).first()
        if department is None:
            messages.error(request, 'Selected department was not found.')
            return redirect(next_url)
        user.department = department
        dept_label = department.name
    else:
        user.department = None
        dept_label = 'No department'

    if hasattr(user, 'updated_by'):
        user.updated_by = request.user
        user.save(update_fields=['department', 'updated_by'])
    else:
        user.save(update_fields=['department'])

    messages.success(request, f"Assigned {user.username} to {dept_label}.")
    return redirect(next_url)


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_deactivation_audit_list(request):
    query = (request.GET.get('q') or '').strip()
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()
    
    start_date = None
    end_date = None
    
    # Parse dates
    if start_date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if end_date_str:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    audits = _filtered_deactivation_audits(query, start_date, end_date)
    return render(
        request,
        'accounts/user_deactivation_audit_list.html',
        {
            'audits': audits,
            'query': query,
            'start_date': start_date_str,
            'end_date': end_date_str,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_deactivation_audit_detail(request, pk):
    audit = get_object_or_404(UserDeactivationAudit.objects.select_related('user', 'admin'), pk=pk)
    return render(
        request,
        'accounts/user_deactivation_audit_detail.html',
        {
            'audit': audit,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_archive_list(request):
    query = (request.GET.get('q') or '').strip()
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()
    
    start_date = None
    end_date = None
    
    # Parse dates
    if start_date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if end_date_str:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    archives = _filtered_user_archives(query, start_date, end_date)
    return render(
        request,
        'accounts/user_archive_list.html',
        {
            'archives': archives,
            'query': query,
            'start_date': start_date_str,
            'end_date': end_date_str,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_archive_detail(request, pk):
    archive = get_object_or_404(UserArchive.objects.select_related('archived_by'), pk=pk)
    return render(
        request,
        'accounts/user_archive_detail.html',
        {
            'archive': archive,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_deactivation_audit_export_excel(request):
    query = (request.GET.get('q') or '').strip()
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()
    
    start_date = None
    end_date = None
    
    # Parse dates
    if start_date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if end_date_str:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    audits = _filtered_deactivation_audits(query, start_date, end_date)

    headers = [
        'User',
        'Employee ID',
        'Admin',
        'Deactivated At',
        'Systems Affected',
        'Hardware Affected',
        'Hardware Action',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Deactivation Audit'
    ws.append(headers)
    for audit in audits:
        ws.append([
            audit.user_full_name or audit.user_username,
            audit.user_employee_id or '',
            audit.admin.get_full_name() if audit.admin else (audit.admin.username if audit.admin else ''),
            timezone.localtime(audit.deactivated_at).strftime('%Y-%m-%d %H:%M'),
            len(audit.system_assignments or []),
            len(audit.hardware_assignments or []),
            audit.hardware_status_action,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename_suffix = f"_q_{query}" if query else ""
    if start_date_str and end_date_str:
        filename_suffix += f"_{start_date_str}_to_{end_date_str}"
    response['Content-Disposition'] = f'attachment; filename="user_deactivation_audit{filename_suffix}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_deactivation_audit_export_pdf(request):
    query = (request.GET.get('q') or '').strip()
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()
    
    start_date = None
    end_date = None
    
    # Parse dates
    if start_date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if end_date_str:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    audits = _filtered_deactivation_audits(query, start_date, end_date)

    data = [['User', 'Employee ID', 'Admin', 'Deactivated At', 'Systems', 'Hardware', 'Hardware Action']]
    for audit in audits:
        data.append([
            audit.user_full_name or audit.user_username,
            audit.user_employee_id or '',
            audit.admin.get_full_name() if audit.admin else (audit.admin.username if audit.admin else ''),
            timezone.localtime(audit.deactivated_at).strftime('%Y-%m-%d %H:%M'),
            str(len(audit.system_assignments or [])),
            str(len(audit.hardware_assignments or [])),
            audit.hardware_status_action,
        ])

    buffer_response = HttpResponse(content_type='application/pdf')
    filename_suffix = f"_q_{query}" if query else ""
    buffer_response['Content-Disposition'] = f'attachment; filename="user_deactivation_audit{filename_suffix}.pdf"'

    page_size = landscape(A4)
    c = canvas.Canvas(buffer_response, pagesize=page_size)
    width, height = page_size
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    available_width = width - 40
    available_height = height - 60
    table.wrapOn(c, available_width, available_height)
    table.drawOn(c, 20, height - 40 - table._height)
    c.showPage()
    c.save()
    return buffer_response


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_archive_export_excel(request):
    query = (request.GET.get('q') or '').strip()
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()
    
    start_date = None
    end_date = None
    
    # Parse dates
    if start_date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if end_date_str:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    archives = _filtered_user_archives(query, start_date, end_date)

    headers = ['Username', 'Full Name', 'Employee ID', 'Department', 'Archived By', 'Archived At']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Archived Users'
    ws.append(headers)
    for record in archives:
        ws.append([
            record.username,
            record.full_name or '',
            record.employee_id or '',
            record.department_name or '',
            record.archived_by.get_full_name() if record.archived_by else (record.archived_by.username if record.archived_by else ''),
            timezone.localtime(record.archived_at).strftime('%Y-%m-%d %H:%M'),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename_suffix = f"_q_{query}" if query else ""
    if start_date_str and end_date_str:
        filename_suffix += f"_{start_date_str}_to_{end_date_str}"
    response['Content-Disposition'] = f'attachment; filename="archived_users{filename_suffix}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def user_archive_export_pdf(request):
    query = (request.GET.get('q') or '').strip()
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()
    
    start_date = None
    end_date = None
    
    # Parse dates
    if start_date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if end_date_str:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    archives = _filtered_user_archives(query, start_date, end_date)

    data = [['Username', 'Full Name', 'Employee ID', 'Department', 'Archived By', 'Archived At']]
    for record in archives:
        data.append([
            record.username,
            record.full_name or '',
            record.employee_id or '',
            record.department_name or '',
            record.archived_by.get_full_name() if record.archived_by else (record.archived_by.username if record.archived_by else ''),
            timezone.localtime(record.archived_at).strftime('%Y-%m-%d %H:%M'),
        ])

    buffer_response = HttpResponse(content_type='application/pdf')
    filename_suffix = f"_q_{query}" if query else ""
    if start_date_str and end_date_str:
        filename_suffix += f"_{start_date_str}_to_{end_date_str}"
    buffer_response['Content-Disposition'] = f'attachment; filename="archived_users{filename_suffix}.pdf"'

    page_size = landscape(A4)
    c = canvas.Canvas(buffer_response, pagesize=page_size)
    width, height = page_size
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    available_width = width - 40
    available_height = height - 60
    table.wrapOn(c, available_width, available_height)
    table.drawOn(c, 20, height - 40 - table._height)
    c.showPage()
    c.save()
    return buffer_response
