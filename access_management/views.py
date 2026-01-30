import csv
from datetime import datetime, timedelta
from io import BytesIO, StringIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer

from .models import (
    UserSystemAccess,
    AccessHistory,
    QuarterlyAccessReview,
    PermissionChangeDocumentation,
    QuarterlyActiveUserReview,
    MonthlyObsoleteAccountReview,
    AccessRemovalDocumentation,
)
from .forms import (
    QuarterlyAccessReviewForm,
    PermissionChangeDocumentationForm,
    BulkQuarterlyReviewForm,
    QuarterlyActiveUserReviewForm,
    MonthlyObsoleteAccountReviewForm,
    AccessRemovalDocumentationForm,
    get_current_quarter_label,
)
from .reporting import build_policy_drift_snapshot, generate_policy_drift_rows
from .utils import (
    is_generic_username,
    detect_generic_accounts,
    get_generic_accounts_by_system,
    get_unremediated_generic_accounts,
    identify_obsolete_accounts,
    get_unapproved_access_records,
)
from accounts.models import CustomUser
from systems.models import System, SystemContract, SystemSubscriptionTier
from departments.models import Department


def _format_datetime(value):
    if not value:
        return ''
    try:
        return timezone.localtime(value).strftime('%Y-%m-%d %H:%M')
    except (ValueError, TypeError):
        return value.strftime('%Y-%m-%d %H:%M') if hasattr(value, 'strftime') else ''


def _format_datetime_for_input(value):
    if not value:
        return ''
    try:
        return timezone.localtime(value).strftime('%Y-%m-%dT%H:%M')
    except (ValueError, TypeError):
        return value.strftime('%Y-%m-%dT%H:%M') if hasattr(value, 'strftime') else ''


def _parse_datetime_input(value):
    if not value:
        return None
    try:
        parsed = timezone.datetime.fromisoformat(value)
    except (ValueError, TypeError):
        return None
    return parsed


def _get_subscription_tier_options():
    """
    Provide a flat list of subscription tier metadata for assignment forms.
    Used to drive dropdowns filtered per-system on the client.
    """
    contracts = SystemContract.objects.select_related("system").prefetch_related(
        "subscription_tiers"
    )
    options = []
    for contract in contracts:
        currency = contract.contract_fee_currency or ""
        for tier in contract.subscription_tiers.all():
            options.append(
                {
                    "id": tier.id,
                    "system_id": contract.system_id,
                    "system_name": contract.system.name,
                    "name": tier.name,
                    "license_category": tier.license_category or tier.name,
                    "unit_price": tier.unit_price,
                    "billing_frequency": tier.billing_frequency,
                    "currency": currency,
                }
            )
    return options


def _match_subscription_tier_for_system(system_id, license_category):
    """
    Find a subscription tier id for the given system that matches a license category.
    """
    if not license_category:
        return None
    return (
        SystemSubscriptionTier.objects.filter(
            contract__system_id=system_id, license_category__iexact=license_category
        )
        .values_list("id", flat=True)
        .first()
    )


def export_access_assignments_to_excel(queryset):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Access Assignments"

    headers = [
        "User",
        "Username",
        "Account Status",
        "Employment Status",
        "Department",
        "System",
        "System Code",
        "Access Username",
        "Access Type",
        "Access Level",
        "Status",
        "Priority",
        "Request Date",
        "Access Start",
        "Access End",
        "Approved By",
        "Approval Date",
    ]
    worksheet.append(headers)

    for assignment in queryset:
        user = assignment.user
        system = assignment.system
        approved_by = assignment.approved_by

        worksheet.append([
            user.get_full_name() if user else '',
            user.username if user else '',
            "Active" if user and user.is_active else "Inactive" if user else '',
            user.employment_status if user else '',
            user.department.name if user and user.department else '',
            system.name if system else '',
            system.code if system else '',
            assignment.access_username or '',
            assignment.access_type or '',
            assignment.granted_access_level or '',
            assignment.status or '',
            assignment.priority or '',
            _format_datetime(assignment.request_date),
            _format_datetime(assignment.access_start_date),
            _format_datetime(assignment.access_end_date),
            approved_by.get_full_name() if approved_by else '',
            _format_datetime(assignment.approval_date),
        ])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="access_assignments.xlsx"'
    return response


def export_access_assignments_to_pdf(queryset):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    headers = [
        "User",
        "Account Status",
        "Employment Status",
        "System",
        "Access Type",
        "Access Level",
        "Status",
        "Priority",
        "Request Date",
        "Access Period",
        "Approved By",
    ]

    data = [headers]

    for assignment in queryset:
        user = assignment.user
        system = assignment.system
        approved_by = assignment.approved_by

        access_period = _format_datetime(assignment.access_start_date)
        if assignment.access_end_date:
            access_period = f"{access_period} → {_format_datetime(assignment.access_end_date)}" if access_period else _format_datetime(assignment.access_end_date)

        data.append([
            f"{user.get_full_name()} ({user.username})" if user else '',
            "Active" if user and user.is_active else "Inactive" if user else '',
            user.employment_status if user else '',
            f"{system.name} ({system.code})" if system else '',
            assignment.access_type or '',
            assignment.granted_access_level or '',
            assignment.status or '',
            assignment.priority or '',
            _format_datetime(assignment.request_date),
            access_period or '',
            approved_by.get_full_name() if approved_by else '',
        ])

    table = Table(data, repeatRows=1, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LINEBEFORE', (0, 0), (0, -1), 0.25, colors.grey),
        ('LINEAFTER', (-1, 0), (-1, -1), 0.25, colors.grey),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    doc.build([table])
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="access_assignments.pdf"'
    return response


def export_policy_drift_rows_to_csv(rows):
    headers = [
        "Issue Type",
        "User",
        "User Login",
        "Department",
        "System",
        "System Code",
        "External Username",
        "Status",
        "Last Review",
        "Next Review",
        "Detail",
        "Assignment ID",
    ]
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for row in rows:
        writer.writerow([
            row["issue_type"],
            row["user_name"],
            row["user_username"],
            row["department"],
            row["system_name"],
            row["system_code"],
            row["external_username"],
            row["status"],
            row["last_review"],
            row["next_review"],
            row["detail"],
            row["assignment_id"],
        ])

    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="policy_drift_snapshot.csv"'
    return response


def export_policy_drift_rows_to_pdf(rows, summary, stale_threshold_days):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    summary_data = [
        ["Metric", "Count"],
        ["Missing Usernames", summary["missing_usernames"]],
        ["Stale Reviews", summary["stale_reviews"]],
        ["Overlapping Usernames", summary["overlapping_usernames"]],
        ["Assignments Scanned", summary["total_assignments"]],
        ["Threshold (days)", stale_threshold_days],
    ]
    summary_table = Table(summary_data, hAlign='LEFT')
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    details_headers = [
        "Issue",
        "User",
        "System",
        "External Username",
        "Status",
        "Last Review",
        "Next Review",
        "Detail",
    ]
    data = [details_headers]
    for row in rows[:200]:
        system_label = f"{row['system_name']} ({row['system_code']})" if row['system_code'] else row['system_name']
        user_label = f"{row['user_name']} ({row['user_username']})" if row['user_username'] else row['user_name']
        data.append([
            row['issue_type'],
            user_label,
            system_label,
            row['external_username'],
            row['status'],
            row['last_review'],
            row['next_review'],
            row['detail'],
        ])

    details_table = Table(data, repeatRows=1, hAlign='LEFT')
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))

    doc.build([summary_table, Spacer(1, 12), details_table])
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="policy_drift_snapshot.pdf"'
    return response


def export_policy_drift_rows_to_excel(rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Policy Drift Snapshot"

    headers = [
        "Issue Type",
        "User",
        "User Login",
        "Department",
        "System",
        "System Code",
        "External Username",
        "Status",
        "Last Review",
        "Next Review",
        "Detail",
        "Assignment ID",
    ]
    worksheet.append(headers)

    for row in rows:
        worksheet.append([
            row["issue_type"],
            row["user_name"],
            row["user_username"],
            row["department"],
            row["system_name"],
            row["system_code"],
            row["external_username"],
            row["status"],
            row["last_review"],
            row["next_review"],
            row["detail"],
            row["assignment_id"],
        ])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="policy_drift_snapshot.xlsx"'
    return response


def build_cross_system_mapping_rows(assignments, request=None):
    rows = []
    for assignment in assignments:
        user = assignment.user
        system = assignment.system

        artifact_file_url = ''
        if assignment.username_verification_artifact:
            try:
                artifact_file_url = assignment.username_verification_artifact.url
                if request and artifact_file_url:
                    artifact_file_url = request.build_absolute_uri(artifact_file_url)
            except ValueError:
                artifact_file_url = ''

        artifact_external_url = assignment.username_verification_artifact_url or ''
        if request and artifact_external_url and artifact_external_url.startswith('/'):
            artifact_external_url = request.build_absolute_uri(artifact_external_url)

        rows.append({
            "user_name": user.full_name if user else '',
            "user_username": user.username if user else '',
            "employee_id": user.employee_id if user else '',
            "department": user.department.name if user and user.department else '',
            "system_name": system.name if system else '',
            "system_code": system.code if system else '',
            "system_username": assignment.effective_username,
            "access_type": assignment.access_type,
            "status": assignment.status,
            "is_generic": 'Yes' if assignment.is_generic_account else 'No',
            "verified_by": assignment.username_verified_by.full_name if assignment.username_verified_by else '',
            "verified_date": _format_datetime(assignment.username_verified_date),
            "verification_artifact_file": artifact_file_url,
            "verification_artifact_url": artifact_external_url,
            # System Owner / 4.4 metadata
            "system_owner_approved": 'Yes' if assignment.system_owner_approved else 'No',
            "system_owner_name": assignment.system.system_owner.full_name if assignment.system and assignment.system.system_owner else '',
            "system_owner_approval_date": _format_datetime(assignment.system_owner_approval_date),
            "legitimate_business_need": (assignment.legitimate_business_need or '').strip(),
            "business_justification": (assignment.business_justification or '').strip(),
        })
    return rows


def _build_quarter_options(count=8):
    """Return a descending list of quarter labels (e.g., ['2025-Q4', ...])."""
    reference = timezone.now().date()
    current_quarter = ((reference.month - 1) // 3) + 1
    current_year = reference.year
    options = []
    idx = 0
    while len(options) < count:
        quarter = current_quarter - idx
        year = current_year
        while quarter <= 0:
            quarter += 4
            year -= 1
        options.append(f"{year}-Q{quarter}")
        idx += 1
    return options


def _quarter_date_range(label):
    """Return (start_datetime, end_datetime) for a given quarter label."""
    if not label or label == 'all':
        return None, None
    try:
        year_part, quarter_part = label.split('-Q')
        year = int(year_part)
        quarter = int(quarter_part)
        month_lookup = {1: 1, 2: 4, 3: 7, 4: 10}
        start_month = month_lookup[quarter]
    except (ValueError, KeyError):
        return None, None

    start_date = timezone.make_aware(datetime(year, start_month, 1, 0, 0, 0))
    end_month = start_month + 2
    if end_month > 12:
        end_year = year + 1
        end_month = end_month - 12
    else:
        end_year = year
    # Last day of quarter
    if end_month in {1, 3, 5, 7, 8, 10, 12}:
        end_day = 31
    elif end_month == 2:
        end_day = 29 if (end_year % 4 == 0 and (end_year % 100 != 0 or end_year % 400 == 0)) else 28
    else:
        end_day = 30
    end_date = timezone.make_aware(datetime(end_year, end_month, end_day, 23, 59, 59))
    return start_date, end_date


def _annual_review_progress(year=None):
    """Return progress metrics for annual user review coverage."""
    now = timezone.now()
    year = year or now.year
    total_users = CustomUser.objects.included_in_metrics().filter(is_active=True).count()
    reviewed_users = (
        QuarterlyAccessReview.objects.filter(review_date__year=year)
        .values_list('reviewed_user', flat=True)
        .distinct()
        .count()
    )
    remaining = max(total_users - reviewed_users, 0)
    percentage = 100.0 if total_users == 0 else round((reviewed_users / total_users) * 100, 1)
    current_quarter = ((now.month - 1) // 3) + 1
    target_percentage = round(min(100.0, (current_quarter / 4) * 100), 1)
    return {
        'year': year,
        'total_users': total_users,
        'reviewed_users': reviewed_users,
        'remaining_users': remaining,
        'percentage': percentage,
        'target_percentage': target_percentage,
        'on_track': total_users == 0 or percentage >= target_percentage or remaining == 0,
        'complete': remaining == 0,
    }


def _select_assignments_for_bulk(system, users_qty, review_quarter):
    """Return a prioritized list of assignments eligible for the requested quarter."""
    if not system or not users_qty or users_qty <= 0:
        return []

    queryset = (
        UserSystemAccess.objects.select_related('user', 'system')
        .filter(
            system=system,
            status__in=['Active', 'Approved'],
            user__is_active=True,
        )
        .exclude(quarterly_reviews__review_quarter=review_quarter)
    )

    assignments = list(queryset)
    far_future = timezone.now() + timedelta(days=365 * 5)
    assignments.sort(
        key=lambda a: (
            a.next_review_date is None,
            a.next_review_date or far_future,
            (a.user.last_name or '').lower() if a.user else '',
            (a.user.first_name or '').lower() if a.user else '',
        )
    )
    return assignments[:users_qty]


def _update_assignment_review_schedule(assignment, review_date):
    """Update last/next review dates on the underlying assignment."""
    default_frequency = assignment.review_frequency_days or 90
    assignment.last_review_date = review_date
    assignment.next_review_date = review_date + timedelta(days=default_frequency)
    assignment.save(update_fields=['last_review_date', 'next_review_date'])


def _default_permission_label(assignment):
    """Best-effort label describing the approved/actual permission set."""
    return (
        assignment.granted_access_level
        or assignment.access_type
        or assignment.system_username
        or assignment.status
        or 'Not Provided'
    )


def export_quarterly_reviews_to_csv(queryset):
    headers = [
        "Quarter",
        "User",
        "System",
        "Reviewed By",
        "Review Date",
        "Approved Permissions",
        "Actual Permissions",
        "Matches Approved",
        "Discrepancies",
        "System Owner",
        "Owner Confirmed",
        "Owner Confirmation Date",
        "Review Completed",
    ]
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for review in queryset:
        writer.writerow([
            review.review_quarter,
            review.reviewed_user.get_full_name() if review.reviewed_user else '',
            review.system.name if review.system else '',
            review.reviewed_by.get_full_name() if review.reviewed_by else '',
            _format_datetime(review.review_date),
            (review.approved_permissions or '').strip(),
            (review.actual_permissions_in_external_system or '').strip(),
            'Yes' if review.matches_approved else 'No',
            (review.discrepancies or '').strip(),
            review.system_owner.get_full_name() if review.system_owner else '',
            'Yes' if review.system_owner_confirmed else 'No',
            _format_datetime(review.system_owner_confirmed_date),
            'Yes' if review.review_completed else 'No',
        ])

    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="quarterly_access_reviews.csv"'
    return response


def export_cross_system_mapping_to_csv(rows):
    headers = [
        "User",
        "Username",
        "Employee ID",
        "Department",
        "System",
        "System Code",
        "System Username",
        "Access Type",
        "Status",
        "Generic Account",
        "Verified By",
        "Verified Date",
        "Verification Artifact (File)",
        "Verification Artifact (URL)",
        "System Owner",
        "System Owner Approved",
        "System Owner Approval Date",
        "Business Justification",
        "Legitimate Business Need",
    ]
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for row in rows:
        writer.writerow([
            row["user_name"],
            row["user_username"],
            row["employee_id"],
            row["department"],
            row["system_name"],
            row["system_code"],
            row["system_username"],
            row["access_type"],
            row["status"],
            row["is_generic"],
            row["verified_by"],
            row["verified_date"],
            row["verification_artifact_file"],
            row["verification_artifact_url"],
            row["system_owner_name"],
            row["system_owner_approved"],
            row["system_owner_approval_date"],
            row["business_justification"],
            row["legitimate_business_need"],
        ])

    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="cross_system_account_mapping.csv"'
    return response


def export_cross_system_mapping_to_excel(rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cross-System Mapping"

    headers = [
        "User",
        "Username",
        "Employee ID",
        "Department",
        "System",
        "System Code",
        "System Username",
        "Access Type",
        "Status",
        "Generic Account",
        "Verified By",
        "Verified Date",
        "Verification Artifact (File)",
        "Verification Artifact (URL)",
        "System Owner",
        "System Owner Approved",
        "System Owner Approval Date",
        "Business Justification",
        "Legitimate Business Need",
    ]
    worksheet.append(headers)

    for row in rows:
        worksheet.append([
            row["user_name"],
            row["user_username"],
            row["employee_id"],
            row["department"],
            row["system_name"],
            row["system_code"],
            row["system_username"],
            row["access_type"],
            row["status"],
            row["is_generic"],
            row["verified_by"],
            row["verified_date"],
            row["verification_artifact_file"],
            row["verification_artifact_url"],
            row["system_owner_name"],
            row["system_owner_approved"],
            row["system_owner_approval_date"],
            row["business_justification"],
            row["legitimate_business_need"],
        ])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cross_system_account_mapping.xlsx"'
    return response


def export_admin_accounts_to_excel(queryset):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Admin Accounts"

    headers = [
        "User",
        "Username",
        "Department",
        "IT Administrator",
        "System",
        "System Code",
        "Access Type",
        "Admin Access",
        "Admin Account Username",
        "Regular Account Username",
        "Workstation Login",
        "Has Domain Admin",
        "Password Storage Location",
        "Password Stored/Verified Date",
        "Status",
    ]
    worksheet.append(headers)

    for assignment in queryset.select_related("user", "system"):
        user = assignment.user
        system = assignment.system
        worksheet.append([
            user.full_name if user else "",
            user.username if user else "",
            user.department.name if user and user.department else "",
            "Yes" if user and getattr(user, "is_it_administrator", False) else "No",
            system.name if system else "",
            system.code if system else "",
            assignment.access_type,
            "Yes" if assignment.is_admin_access else "",
            assignment.admin_account_username or "",
            assignment.regular_account_username or "",
            "Yes" if assignment.is_workstation_login else "",
            "Yes" if assignment.has_domain_admin else "",
            assignment.admin_password_storage_location or "",
            _format_datetime(assignment.admin_password_stored_date),
            assignment.status,
        ])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="admin_accounts_compliance.xlsx"'
    return response


def export_admin_accounts_to_csv(queryset):
    headers = [
        "User",
        "Username",
        "Department",
        "IT Administrator",
        "System",
        "System Code",
        "Access Type",
        "Admin Access",
        "Admin Account Username",
        "Regular Account Username",
        "Workstation Login",
        "Has Domain Admin",
        "Password Storage Location",
        "Password Stored/Verified Date",
        "Status",
    ]
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for assignment in queryset.select_related("user", "system"):
        user = assignment.user
        system = assignment.system
        writer.writerow([
            user.full_name if user else "",
            user.username if user else "",
            user.department.name if user and user.department else "",
            "Yes" if user and getattr(user, "is_it_administrator", False) else "No",
            system.name if system else "",
            system.code if system else "",
            assignment.access_type,
            "Yes" if assignment.is_admin_access else "",
            assignment.admin_account_username or "",
            assignment.regular_account_username or "",
            "Yes" if assignment.is_workstation_login else "",
            "Yes" if assignment.has_domain_admin else "",
            assignment.admin_password_storage_location or "",
            _format_datetime(assignment.admin_password_stored_date),
            assignment.status,
        ])

    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="admin_accounts_compliance.csv"'
    return response


#@login_required
# def access_assignment_list(request):
#     """List all access assignments with filtering and search"""
    
#     # Get filter parameters
#     status_filter = request.GET.get('status', '')
#     priority_filter = request.GET.get('priority', '')
#     access_type_filter = request.GET.get('access_type', '')
#     system_filter = request.GET.get('system', '')
#     user_filter = request.GET.get('user', '')
#     employment_status_filter = request.GET.get('employment_status', '')
#     user_active_filter = request.GET.get('user_active', '')
#     search_query = request.GET.get('search', '')
    
#     # Base queryset
#     queryset = UserSystemAccess.objects.select_related('user', 'system', 'approved_by').all()
    
#     # Apply filters
#     if status_filter:
#         queryset = queryset.filter(status=status_filter)
#     if priority_filter:
#         queryset = queryset.filter(priority=priority_filter)
#     if access_type_filter:
#         queryset = queryset.filter(access_type=access_type_filter)
#     if system_filter:
#         queryset = queryset.filter(system_id=system_filter)
#     if user_filter:
#         queryset = queryset.filter(user_id=user_filter)
#     if employment_status_filter:
#         queryset = queryset.filter(user__employment_status=employment_status_filter)
#     if user_active_filter == 'active':
#         queryset = queryset.filter(user__is_active=True)
#     elif user_active_filter == 'inactive':
#         queryset = queryset.filter(user__is_active=False)
#     if search_query:
#         queryset = queryset.filter(
#             Q(user__username__icontains=search_query) |
#             Q(user__first_name__icontains=search_query) |
#             Q(user__last_name__icontains=search_query) |
#             Q(system__name__icontains=search_query) |
#             Q(business_justification__icontains=search_query)
#         )
    
#     metrics_queryset = queryset

#     summary_metrics = {
#         'total': metrics_queryset.count(),
#         'active': metrics_queryset.filter(status='Active').count(),
#         'pending': metrics_queryset.filter(status='Pending').count(),
#         'expired': metrics_queryset.filter(status='Expired').count(),
#         'unique_users': metrics_queryset.values('user_id').distinct().count(),
#         'unique_systems': metrics_queryset.values('system_id').distinct().count(),
#     }

#     export_format = request.GET.get('export')
#     if export_format in {'xlsx', 'pdf'}:
#         export_queryset = queryset.order_by('user__first_name', 'user__last_name', '-request_date', 'system__name')
#         if export_format == 'xlsx':
#             return export_access_assignments_to_excel(export_queryset)
#         return export_access_assignments_to_pdf(export_queryset)
    
#     # Pagination
#     queryset = queryset.order_by('user__first_name', 'user__last_name', '-request_date', 'system__name')

#     paginator = Paginator(queryset, 25)
#     page_number = request.GET.get('page')
#     access_assignments = paginator.get_page(page_number)
    
#     # Get filter options
#     systems = System.objects.all().order_by('name')
#     users = CustomUser.objects.all().order_by('first_name', 'last_name')

#     query_params = request.GET.copy()
#     query_params.pop('page', None)
#     query_params.pop('export', None)
#     current_query = query_params.urlencode()

#     context = {
#         'access_assignments': access_assignments,
#         'status_choices': UserSystemAccess.STATUS_CHOICES,
#         'priority_choices': UserSystemAccess.PRIORITY_CHOICES,
#         'access_type_choices': UserSystemAccess.ACCESS_TYPE_CHOICES,
#         'employment_status_choices': CustomUser.EMPLOYMENT_STATUS_CHOICES,
#         'systems': systems,
#         'users': users,
#         'filters': {
#             'status': status_filter,
#             'priority': priority_filter,
#             'access_type': access_type_filter,
#             'system': system_filter,
#             'user': user_filter,
#             'employment_status': employment_status_filter,
#             'user_active': user_active_filter,
#             'search': search_query,
#         },
#         'current_query': current_query,
#         'summary_metrics': summary_metrics,
#     }
    
#     return render(request, 'access_management/access_assignment_list.html', context)
# Enhanced access_assignment_list view with additional features
# Replace your existing access_assignment_list function with this corrected version
# This fixes the syntax error and adds all necessary features

@login_required
def access_assignment_list(request):
    """
    Enhanced list view for access assignments with filtering, search, and metrics
    """
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    access_type_filter = request.GET.get('access_type', '')
    system_filter = request.GET.get('system', '')
    user_filter = request.GET.get('user', '')
    employment_status_filter = request.GET.get('employment_status', '')
    user_active_filter = request.GET.get('user_active', '')
    search_query = request.GET.get('search', '')
    
    # NEW: Additional filters
    days_to_expiry = request.GET.get('days_to_expiry', '')
    days_pending = request.GET.get('days_pending', '')
    page_size = request.GET.get('page_size', '25')
    
    # Base queryset with optimized select_related
    queryset = UserSystemAccess.objects.select_related(
        'user',
        'user__department', 
        'system', 
        'approved_by',
    ).all()
    
    # Apply existing filters
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if priority_filter:
        queryset = queryset.filter(priority=priority_filter)
    if access_type_filter:
        queryset = queryset.filter(access_type=access_type_filter)
    if system_filter:
        queryset = queryset.filter(system_id=system_filter)
    if user_filter:
        queryset = queryset.filter(user_id=user_filter)
    if employment_status_filter:
        queryset = queryset.filter(user__employment_status=employment_status_filter)
    if user_active_filter == 'active':
        queryset = queryset.filter(user__is_active=True)
    elif user_active_filter == 'inactive':
        queryset = queryset.filter(user__is_active=False)
    if search_query:
        queryset = queryset.filter(
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(system__name__icontains=search_query) |
            Q(system__code__icontains=search_query) |
            Q(business_justification__icontains=search_query)
        )
    
    # NEW: Expiry filter - find assignments expiring within X days
    if days_to_expiry:
        try:
            days = int(days_to_expiry)
            expiry_threshold = timezone.now().date() + timedelta(days=days)
            queryset = queryset.filter(
                expiry_date__lte=expiry_threshold,
                expiry_date__gte=timezone.now().date(),
                status='Active'
            )
        except ValueError:
            pass
    
    # NEW: Pending duration filter - find assignments pending for more than X days
    if days_pending:
        try:
            days = int(days_pending)
            pending_threshold = timezone.now() - timedelta(days=days)
            queryset = queryset.filter(
                status='Pending',
                request_date__lte=pending_threshold
            )
        except ValueError:
            pass
    
    # Store queryset for metrics before pagination
    metrics_queryset = queryset
    
    # Calculate summary metrics
    summary_metrics = {
        'total': metrics_queryset.count(),
        'active': metrics_queryset.filter(status='Active').count(),
        'pending': metrics_queryset.filter(status='Pending').count(),
        'expired': metrics_queryset.filter(status='Expired').count(),
        'unique_users': metrics_queryset.values('user_id').distinct().count(),
        'unique_systems': metrics_queryset.values('system_id').distinct().count(),
    }
    
    # NEW: Calculate quick filter counts for the chips
    now = timezone.now()
    expiring_threshold = now.date() + timedelta(days=30)
    pending_threshold = now - timedelta(days=7)
    
    # Quick filter counts with safe calculations
    quick_filter_counts = {
        'my_pending': 0,
        'expiring_soon': 0,
        'high_priority': 0,
        'pending_over_7_days': 0,
    }
    
    # Calculate expiring soon count
    try:
        quick_filter_counts['expiring_soon'] = metrics_queryset.filter(
            expiry_date__lte=expiring_threshold,
            expiry_date__gte=now.date(),
            status='Active'
        ).count()
    except Exception:
        pass
    
    # Calculate high priority count
    try:
        quick_filter_counts['high_priority'] = metrics_queryset.filter(
            priority__in=['High', 'Critical']
        ).count()
    except Exception:
        pass
    
    # Calculate pending over 7 days count
    try:
        quick_filter_counts['pending_over_7_days'] = metrics_queryset.filter(
            status='Pending',
            request_date__lte=pending_threshold
        ).count()
    except Exception:
        pass
    
    # Calculate my pending approvals count if user is authenticated
    if request.user.is_authenticated:
        try:
            quick_filter_counts['my_pending'] = metrics_queryset.filter(
                status='Pending'
            ).filter(
                Q(approved_by=request.user) | Q(system_owner_approver=request.user)
            ).count()
        except Exception:
            pass
    
    # Handle export requests
    export_format = request.GET.get('export')
    if export_format in {'xlsx', 'pdf'}:
        export_queryset = queryset.order_by(
            'user__first_name', 
            'user__last_name', 
            '-request_date', 
            'system__name'
        )
        if export_format == 'xlsx':
            return export_access_assignments_to_excel(export_queryset)
        return export_access_assignments_to_pdf(export_queryset)
    
    # NEW: Dynamic page size with validation
    try:
        page_size = int(page_size)
        if page_size not in [10, 25, 50, 100]:
            page_size = 25
    except (ValueError, TypeError):
        page_size = 25
    
    # Order queryset
    queryset = queryset.order_by(
        'user__first_name', 
        'user__last_name', 
        '-request_date', 
        'system__name'
    )
    
    # Pagination
    paginator = Paginator(queryset, page_size)
    page_number = request.GET.get('page')
    access_assignments = paginator.get_page(page_number)
    
    # NEW: Annotate assignments with computed fields for better UX
    for assignment in access_assignments:
        # Calculate days until expiry
        if assignment.expiry_date:
            try:
                delta = assignment.expiry_date - now.date()
                assignment.days_until_expiry = delta.days
                assignment.is_expiring_soon = 0 <= assignment.days_until_expiry <= 30
            except Exception:
                assignment.days_until_expiry = None
                assignment.is_expiring_soon = False
        else:
            assignment.days_until_expiry = None
            assignment.is_expiring_soon = False
        
        # Calculate pending days for pending assignments
        if assignment.status == 'Pending' and assignment.request_date:
            try:
                delta = now - assignment.request_date
                assignment.pending_days = delta.days
            except Exception:
                assignment.pending_days = 0
        else:
            assignment.pending_days = 0
        
        # Check if current user can edit this assignment
        assignment.can_edit = (
            request.user.is_superuser or 
            request.user == assignment.user or
            request.user.has_perm('access_management.change_usersystemaccess')
        )
    
    # Get filter options for dropdowns
    systems = System.objects.all().order_by('name')
    users = CustomUser.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    # Build current query string for pagination links
    query_params = request.GET.copy()
    query_params.pop('page', None)
    query_params.pop('export', None)
    current_query = query_params.urlencode()
    
    # Build context
    context = {
        'access_assignments': access_assignments,
        'status_choices': UserSystemAccess.STATUS_CHOICES,
        'priority_choices': UserSystemAccess.PRIORITY_CHOICES,
        'access_type_choices': UserSystemAccess.ACCESS_TYPE_CHOICES,
        'employment_status_choices': CustomUser.EMPLOYMENT_STATUS_CHOICES,
        'systems': systems,
        'users': users,
        'filters': {
            'status': status_filter,
            'priority': priority_filter,
            'access_type': access_type_filter,
            'system': system_filter,
            'user': user_filter,
            'employment_status': employment_status_filter,
            'user_active': user_active_filter,
            'search': search_query,
            'days_to_expiry': days_to_expiry,
            'days_pending': days_pending,
        },
        'current_query': current_query,
        'summary_metrics': summary_metrics,
        'quick_filter_counts': quick_filter_counts,
        'page_size': page_size,
    }
    
    return render(request, 'access_management/access_assignment_list.html', context)

# Optional: Add this helper view for AJAX live search
@login_required
def access_assignment_search_ajax(request):
    """
    AJAX endpoint for live search autocomplete
    Returns JSON with matching assignments
    """
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    assignments = UserSystemAccess.objects.select_related(
        'user', 'system'
    ).filter(
        Q(user__username__icontains=query) |
        Q(user__first_name__icontains=query) |
        Q(user__last_name__icontains=query) |
        Q(system__name__icontains=query)
    )[:10]
    
    results = [{
        'id': a.id,
        'user': a.user.get_full_name(),
        'username': a.user.username,
        'system': a.system.name,
        'status': a.status,
        'url': f'/access-management/assignments/{a.id}/'
    } for a in assignments]
    
    return JsonResponse({'results': results})


# Optional: Add bulk approval endpoint
@login_required
@require_http_methods(["POST"])
def bulk_approve_assignments(request):
    """
    Bulk approve multiple assignments at once
    """
    if not request.user.has_perm('access_management.can_approve'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    assignment_ids = request.POST.getlist('assignment_ids[]')
    comments = request.POST.get('comments', '')
    
    approved_count = 0
    errors = []
    
    for assignment_id in assignment_ids:
        try:
            assignment = UserSystemAccess.objects.get(id=assignment_id, status='Pending')
            
            # Perform approval logic
            assignment.status = 'Approved'
            assignment.approved_by = request.user
            assignment.approval_date = timezone.now()
            assignment.approval_comments = comments
            assignment.save()
            
            # Log history
            AccessHistory.objects.create(
                access_assignment=assignment,
                action='Approved',
                performed_by=request.user,
                comments=comments
            )
            
            approved_count += 1
            
        except UserSystemAccess.DoesNotExist:
            errors.append(f'Assignment {assignment_id} not found or not pending')
        except Exception as e:
            errors.append(f'Error approving {assignment_id}: {str(e)}')
    
    return JsonResponse({
        'success': True,
        'approved_count': approved_count,
        'errors': errors
    })


# Optional: Add bulk rejection endpoint
@login_required
@require_http_methods(["POST"])
def bulk_reject_assignments(request):
    """
    Bulk reject multiple assignments at once
    """
    if not request.user.has_perm('access_management.can_approve'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    assignment_ids = request.POST.getlist('assignment_ids[]')
    reason = request.POST.get('reason', '')
    
    if not reason:
        return JsonResponse({'error': 'Rejection reason is required'}, status=400)
    
    rejected_count = 0
    errors = []
    
    for assignment_id in assignment_ids:
        try:
            assignment = UserSystemAccess.objects.get(id=assignment_id, status='Pending')
            
            # Perform rejection logic
            assignment.status = 'Rejected'
            assignment.approved_by = request.user
            assignment.approval_date = timezone.now()
            assignment.rejection_reason = reason
            assignment.save()
            
            # Log history
            AccessHistory.objects.create(
                access_assignment=assignment,
                action='Rejected',
                performed_by=request.user,
                comments=reason
            )
            
            rejected_count += 1
            
        except UserSystemAccess.DoesNotExist:
            errors.append(f'Assignment {assignment_id} not found or not pending')
        except Exception as e:
            errors.append(f'Error rejecting {assignment_id}: {str(e)}')
    
    return JsonResponse({
        'success': True,
        'rejected_count': rejected_count,
        'errors': errors
    })

@login_required
def my_pending_approvals(request):
    """
    Compact queue of access assignments that are currently pending
    and where the logged-in user is involved as approver (IT approver
    or system owner approver).
    """
    user = request.user

    pending_qs = UserSystemAccess.objects.select_related(
        "user", "system"
    ).filter(
        status="Pending"
    ).filter(
        Q(approved_by=user) | Q(system_owner_approver=user)
    ).order_by("-priority", "request_date")

    # Summary metrics for quick triage
    oldest = pending_qs.order_by("request_date").first()
    oldest_days = None
    if oldest and oldest.request_date:
        delta = timezone.now() - oldest.request_date
        oldest_days = max(delta.days, 0)

    summary_metrics = {
        "total": pending_qs.count(),
        "missing_owner": pending_qs.filter(system_owner_approved=False).count(),
        "high_critical": pending_qs.filter(priority__in=["High", "Critical"]).count(),
        "oldest_days": oldest_days,
    }

    context = {
        "pending_assignments": pending_qs,
        "summary_metrics": summary_metrics,
    }
    return render(request, "access_management/my_pending_approvals.html", context)


@login_required
def approval_summary_dashboard(request):
    """
    Summary dashboard of access approvals per system and per department.
    Groups counts of assignments by status to give a quick overview.
    """
    # Aggregation per system
    system_rows = (
        UserSystemAccess.objects.select_related("system")
        .values("system_id", "system__name", "system__code")
        .annotate(
            total=Count("id"),
            pending=Count("id", filter=Q(status="Pending")),
            approved=Count("id", filter=Q(status="Approved")),
            active=Count("id", filter=Q(status="Active")),
            revoked=Count("id", filter=Q(status="Revoked")),
            expired=Count("id", filter=Q(status="Expired")),
        )
        .order_by("system__name")
    )

    # Aggregation per department (based on the user.department relation)
    dept_rows = (
        UserSystemAccess.objects.select_related("user__department")
        .values("user__department_id", "user__department__name")
        .annotate(
            total=Count("id"),
            pending=Count("id", filter=Q(status="Pending")),
            approved=Count("id", filter=Q(status="Approved")),
            active=Count("id", filter=Q(status="Active")),
            revoked=Count("id", filter=Q(status="Revoked")),
            expired=Count("id", filter=Q(status="Expired")),
        )
        .order_by("user__department__name")
    )

    context = {
        "system_rows": system_rows,
        "department_rows": dept_rows,
    }
    return render(request, "access_management/approval_summary_dashboard.html", context)


def _export_unapproved_to_csv(queryset):
    """Export unapproved access records to CSV."""
    output = StringIO()
    writer = csv.writer(output)
    headers = [
        "User",
        "Username",
        "Department",
        "System",
        "System Code",
        "Status",
        "Has IT Approval",
        "Has System Owner Approval",
        "Request Date",
    ]
    writer.writerow(headers)

    for a in queryset.select_related("user__department", "system", "approved_by"):
        user = a.user
        system = a.system
        writer.writerow([
            user.get_full_name() if user else "",
            user.username if user else "",
            user.department.name if user and user.department else "",
            system.name if system else "",
            system.code if system else "",
            a.status,
            "Yes" if a.approved_by else "No",
            "Yes" if a.system_owner_approved else "No",
            _format_datetime(a.request_date),
        ])

    response = HttpResponse(output.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="unapproved_access.csv"'
    return response


def _export_unapproved_to_excel(queryset):
    """Export unapproved access records to Excel."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Unapproved Access"

    headers = [
        "User",
        "Username",
        "Department",
        "System",
        "System Code",
        "Status",
        "Has IT Approval",
        "Has System Owner Approval",
        "Request Date",
    ]
    worksheet.append(headers)

    for a in queryset.select_related("user__department", "system", "approved_by"):
        user = a.user
        system = a.system
        worksheet.append([
            user.get_full_name() if user else "",
            user.username if user else "",
            user.department.name if user and user.department else "",
            system.name if system else "",
            system.code if system else "",
            a.status,
            "Yes" if a.approved_by else "No",
            "Yes" if a.system_owner_approved else "No",
            _format_datetime(a.request_date),
        ])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="unapproved_access.xlsx"'
    return response


@login_required
def unapproved_access_list(request):
    """
    All unapproved access detected by get_unapproved_access_records(),
    with simple filters and CSV/Excel export.
    """
    qs = get_unapproved_access_records()

    # Bulk approval of selected records
    if request.method == "POST":
        selected_ids = request.POST.getlist("assignment_ids")
        bulk_action = request.POST.get("bulk_action") or "it_approve"
        if not selected_ids:
            messages.warning(request, "No assignments selected for bulk approval.")
            return redirect("access_management:unapproved_access_list")

        to_update = qs.filter(id__in=selected_ids)
        now = timezone.now()

        if bulk_action == "owner_approve":
            updated = 0
            for assignment in to_update:
                # Only mark missing system owner approval; do not change status
                if not assignment.system_owner_approved:
                    assignment.system_owner_approved = True
                    assignment.system_owner_approval_date = now
                    assignment.system_owner_approver = request.user
                    assignment.save(update_fields=[
                        "system_owner_approved",
                        "system_owner_approval_date",
                        "system_owner_approver",
                    ])
                    updated += 1
            if updated:
                messages.success(request, f"Bulk marked system owner approval for {updated} assignment(s).")
            else:
                messages.info(request, "Selected assignments already had system owner approval.")
        else:
            # Default: IT approval
            updated = 0
            for assignment in to_update:
                # Only fill in missing IT approval; do not change status
                if assignment.approved_by is None:
                    assignment.approved_by = request.user
                    if not assignment.approval_date:
                        assignment.approval_date = now
                    assignment.save(update_fields=["approved_by", "approval_date"])
                    updated += 1

            if updated:
                messages.success(request, f"Bulk approved IT approval for {updated} assignment(s).")
            else:
                messages.info(request, "Selected assignments already had IT approval.")

        return redirect("access_management:unapproved_access_list")

    system_filter = request.GET.get("system") or ""
    dept_filter = request.GET.get("department") or ""
    gap_type = request.GET.get("gap_type") or ""
    search_query = request.GET.get("search") or ""

    if system_filter:
        qs = qs.filter(system_id=system_filter)
    if dept_filter:
        qs = qs.filter(user__department_id=dept_filter)
    if gap_type == "no_it":
        qs = qs.filter(approved_by__isnull=True)
    elif gap_type == "no_owner":
        qs = qs.filter(system_owner_approved=False)

    if search_query:
        qs = qs.filter(
            Q(user__username__icontains=search_query)
            | Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
            | Q(system__name__icontains=search_query)
            | Q(system__code__icontains=search_query)
        )

    export = request.GET.get("export")
    ordered_qs = qs.order_by("system__name", "user__first_name", "user__last_name")
    if export == "csv":
        return _export_unapproved_to_csv(ordered_qs)
    if export == "xlsx":
        return _export_unapproved_to_excel(ordered_qs)

    systems = System.objects.all().order_by("name")
    departments = Department.objects.all().order_by("name")

    query_params = request.GET.copy()
    query_params.pop("export", None)
    current_query = query_params.urlencode()

    context = {
        "assignments": ordered_qs,
        "systems": systems,
        "departments": departments,
        "filters": {
            "system": system_filter,
            "department": dept_filter,
            "gap_type": gap_type,
            "search": search_query,
        },
        "current_query": current_query,
    }
    return render(request, "access_management/unapproved_access_list.html", context)


@login_required
def my_unapproved_access_gaps(request):
    """
    Unapproved access gaps scoped to systems owned by the current user.
    """
    user = request.user
    # Systems where this user is the system owner; align with System model fields
    owned_systems = System.objects.filter(system_owner=user)

    qs = get_unapproved_access_records().filter(system__in=owned_systems)

    assignments = qs.order_by("system__name", "user__first_name", "user__last_name")

    context = {
        "assignments": assignments,
    }
    return render(request, "access_management/my_unapproved_access_gaps.html", context)


@login_required
def access_assignment_detail(request, pk):
    """Detail view of an access assignment"""
    access_assignment = get_object_or_404(
        UserSystemAccess.objects.select_related('user', 'system', 'approved_by', 'requested_by'),
        pk=pk
    )
    
    # Get recent access history
    access_history = AccessHistory.objects.filter(
        user_system_access=access_assignment
    ).select_related('user', 'system').order_by('-accessed_at')[:10]
    
    context = {
        'access_assignment': access_assignment,
        'access_history': access_history,
    }
    
    return render(request, 'access_management/access_assignment_detail.html', context)


@login_required
def access_assignment_create(request):
    """Create a new access assignment"""
    selected_subscription_tier_id = (
        request.POST.get('subscription_tier', '') if request.method == 'POST' else ''
    )
    license_category_value = (
        request.POST.get('license_category', '').strip() if request.method == 'POST' else ''
    )

    if request.method == 'POST':
        user_id = request.POST.get('user')
        system_id = request.POST.get('system')
        access_type = request.POST.get('access_type')
        request_type = request.POST.get('request_type')
        priority = request.POST.get('priority')
        business_justification = request.POST.get('business_justification')
        legitimate_business_need = request.POST.get('legitimate_business_need', '').strip()
        requested_access_duration = request.POST.get('requested_access_duration')
        technical_requirements = request.POST.get('technical_requirements')
        access_start_date = request.POST.get('access_start_date')
        access_end_date = request.POST.get('access_end_date')
        username_verified_by_id = request.POST.get('username_verified_by')
        username_verified_date_raw = request.POST.get('username_verified_date')
        username_verification_artifact_url = request.POST.get('username_verification_artifact_url', '').strip()
        verification_artifact_file = request.FILES.get('username_verification_artifact')

        # System Owner authorization fields (RHG 4.4)
        system_owner_approved = request.POST.get('system_owner_approved') == 'on'
        system_owner_approval_date_raw = request.POST.get('system_owner_approval_date')
        system_owner_approver_id = request.POST.get('system_owner_approver')
        
        try:
            user = CustomUser.objects.get(id=user_id)
            system = System.objects.get(id=system_id)

            # Resolve subscription tier / license category mapping for subscription-based systems
            tiers_for_system = list(
                SystemSubscriptionTier.objects.select_related('contract').filter(contract__system=system)
            )
            tier_lookup = {str(tier.id): tier for tier in tiers_for_system}
            selected_tier_obj = (
                tier_lookup.get(str(selected_subscription_tier_id)) if selected_subscription_tier_id else None
            )
            if not selected_tier_obj and not license_category_value and len(tiers_for_system) == 1:
                selected_tier_obj = tiers_for_system[0]
                selected_subscription_tier_id = str(selected_tier_obj.id)
            if selected_tier_obj:
                license_category_value = selected_tier_obj.license_category or selected_tier_obj.name

            # Check if access already exists
            if UserSystemAccess.objects.filter(user=user, system=system).exists():
                messages.error(request, f'Access assignment for {user.full_name} to {system.name} already exists.')
                return redirect('access_management:access_assignment_create')
            
            # Get new fields
            system_username = request.POST.get('system_username', '').strip()
            access_username_raw = request.POST.get('access_username')
            if access_username_raw is not None:
                access_username = access_username_raw.strip()
            else:
                access_username = None
            is_generic = request.POST.get('is_generic_account') == 'on'

            # Administrator-equivalent access metadata (RHG 4.3)
            # Default is_admin_access to True when access_type is Admin/Super Admin even if checkbox is not ticked.
            is_admin_access = (
                request.POST.get('is_admin_access') == 'on'
                or access_type in ['Admin', 'Super Admin']
            )
            has_separate_admin_account = request.POST.get('has_separate_admin_account') == 'on'
            admin_account_username = request.POST.get('admin_account_username', '').strip()
            regular_account_username = request.POST.get('regular_account_username', '').strip()
            is_workstation_login = request.POST.get('is_workstation_login') == 'on'
            has_domain_admin = request.POST.get('has_domain_admin') == 'on'
            admin_password_storage_location = request.POST.get('admin_password_storage_location', '').strip()
            admin_password_stored_date_raw = request.POST.get('admin_password_stored_date')
            
            # If system_username is empty but access_username exists, use access_username
            # This helps migrate legacy data
            if not system_username and access_username:
                system_username = access_username
            
            # Convert empty string to None for database consistency
            system_username = system_username if system_username else None
            access_username = access_username if access_username else None
            username_verification_artifact_url = username_verification_artifact_url or None

            username_verified_by = None
            if username_verified_by_id:
                username_verified_by = CustomUser.objects.filter(id=username_verified_by_id).first()
            if not username_verified_by and username_verified_date_raw:
                username_verified_by = request.user

            username_verified_date = _parse_datetime_input(username_verified_date_raw)

            # Parse admin password stored date
            admin_password_stored_date = _parse_datetime_input(admin_password_stored_date_raw)

            # Parse system owner approval date
            system_owner_approval_date = _parse_datetime_input(system_owner_approval_date_raw)

            system_owner_approver = None
            if system_owner_approver_id:
                system_owner_approver = CustomUser.objects.filter(id=system_owner_approver_id).first()
            
            # Create new access assignment
            access_assignment = UserSystemAccess.objects.create(
                user=user,
                system=system,
                access_type=access_type,
                request_type=request_type or 'New Access',
                priority=priority,
                business_justification=business_justification,
                legitimate_business_need=legitimate_business_need or None,
                requested_access_duration=int(requested_access_duration) if requested_access_duration else None,
                technical_requirements=technical_requirements,
                access_start_date=timezone.datetime.fromisoformat(access_start_date) if access_start_date else None,
                access_end_date=timezone.datetime.fromisoformat(access_end_date) if access_end_date else None,
                system_username=system_username,
                access_username=access_username,
                is_generic_account=is_generic,
                # Administrator-equivalent access metadata
                is_admin_access=is_admin_access,
                has_separate_admin_account=has_separate_admin_account,
                admin_account_username=admin_account_username or None,
                regular_account_username=regular_account_username or None,
                is_workstation_login=is_workstation_login,
                has_domain_admin=has_domain_admin,
                admin_password_storage_location=admin_password_storage_location or None,
                admin_password_stored_date=admin_password_stored_date,
                admin_password_stored_by=request.user if admin_password_stored_date else None,
                system_owner_approved=system_owner_approved,
                system_owner_approval_date=system_owner_approval_date,
                system_owner_approver=system_owner_approver,
                license_category=license_category_value or None,
                username_verified_by=username_verified_by,
                username_verified_date=username_verified_date,
                username_verification_artifact=verification_artifact_file,
                username_verification_artifact_url=username_verification_artifact_url,
                requested_by=request.user,
                created_by=request.user,
                updated_by=request.user
            )
            
            # Auto-detect generic accounts
            if system_username:
                access_assignment.mark_as_generic_if_needed()
                access_assignment.save()
                
                # Show warning if generic account detected
                if access_assignment.is_generic_account:
                    messages.warning(
                        request,
                        f'Warning: Username "{system_username}" appears to be a generic account. '
                        f'Please ensure this is remediated per RHG Access Control Policy.'
                    )
            
            # Create access history entry
            AccessHistory.objects.create(
                user=user,
                system=system,
                user_system_access=access_assignment,
                action='Requested',
                action_description=f'Access requested by {request.user.full_name}',
                created_by=request.user
            )
            
            messages.success(request, f'Access assignment created successfully for {user.full_name} to {system.name}.')
            return redirect('access_management:access_assignment_detail', pk=access_assignment.pk)
            
        except (CustomUser.DoesNotExist, System.DoesNotExist):
            messages.error(request, 'Invalid user or system selected.')
        except Exception as e:
            messages.error(request, f'Error creating access assignment: {str(e)}')
    
    # Get data for form
    systems = System.objects.all().order_by('name')
    users = CustomUser.objects.all().order_by('first_name', 'last_name')
    verifiers = users
    default_verifier_id = ''
    if request.user and getattr(request.user, 'pk', None):
        default_verifier_id = str(request.user.pk)
    selected_user_id = (request.POST.get('user') if request.method == 'POST' else None) or ''
    selected_system_id = (request.POST.get('system') if request.method == 'POST' else None) or (request.GET.get('system') or '')
    selected_access_type = (request.POST.get('access_type') if request.method == 'POST' else '') or ''
    selected_request_type = (request.POST.get('request_type') if request.method == 'POST' else '') or ''
    selected_priority = (request.POST.get('priority') if request.method == 'POST' else '') or ''
    
    context = {
        'systems': systems,
        'users': users,
        'verifiers': verifiers,
        'access_type_choices': UserSystemAccess.ACCESS_TYPE_CHOICES,
        'request_type_choices': UserSystemAccess.REQUEST_TYPE_CHOICES,
        'priority_choices': UserSystemAccess.PRIORITY_CHOICES,
        'selected_user_id': str(selected_user_id),
        'selected_system_id': str(selected_system_id),
        'selected_access_type': selected_access_type,
        'selected_request_type': selected_request_type,
        'selected_priority': selected_priority,
        'business_justification_value': request.POST.get('business_justification', ''),
        'requested_access_duration_value': request.POST.get('requested_access_duration', ''),
        'access_start_date_value': request.POST.get('access_start_date', ''),
        'access_end_date_value': request.POST.get('access_end_date', ''),
        'license_category_value': license_category_value,
        'subscription_tiers': _get_subscription_tier_options(),
        'selected_subscription_tier_id': str(selected_subscription_tier_id),
        'access_url_value': request.POST.get('access_url', ''),
        'granted_access_level_value': request.POST.get('granted_access_level', ''),
        'technical_requirements_value': request.POST.get('technical_requirements', ''),
        'security_clearance_required_value': request.POST.get('security_clearance_required', ''),
        'data_access_level_value': request.POST.get('data_access_level', ''),
        'risk_assessment_score_value': request.POST.get('risk_assessment_score', ''),
        'review_frequency_days_value': request.POST.get('review_frequency_days', ''),
        'special_instructions_value': request.POST.get('special_instructions', ''),
        'compliance_requirements_value': request.POST.get('compliance_requirements', ''),
        # System Owner authorization (4.4)
        'system_owner_approved_value': request.POST.get('system_owner_approved', ''),
        'system_owner_approval_date_value': request.POST.get('system_owner_approval_date', ''),
        'system_owner_approver_value': request.POST.get('system_owner_approver', ''),
        'legitimate_business_need_value': request.POST.get('legitimate_business_need', ''),
        'system_username_value': request.POST.get('system_username', ''),
        'username_verified_by_value': request.POST.get('username_verified_by', default_verifier_id),
        'username_verified_date_value': request.POST.get('username_verified_date', ''),
        'username_verification_artifact_url_value': request.POST.get('username_verification_artifact_url', ''),
        # Admin access governance defaults (creation form)
        'is_admin_access_value': request.POST.get('is_admin_access', ''),
        'has_separate_admin_account_value': request.POST.get('has_separate_admin_account', ''),
        'admin_account_username_value': request.POST.get('admin_account_username', ''),
        'regular_account_username_value': request.POST.get('regular_account_username', ''),
        'is_workstation_login_value': request.POST.get('is_workstation_login', ''),
        'has_domain_admin_value': request.POST.get('has_domain_admin', ''),
        'admin_password_storage_location_value': request.POST.get('admin_password_storage_location', ''),
        'admin_password_stored_date_value': request.POST.get('admin_password_stored_date', ''),
    }
    
    return render(request, 'access_management/access_assignment_form.html', context)


@login_required
def access_assignment_update(request, pk):
    """Update an existing access assignment"""
    access_assignment = get_object_or_404(UserSystemAccess, pk=pk)
    selected_subscription_tier_id = (
        request.POST.get('subscription_tier', '')
        if request.method == 'POST'
        else (_match_subscription_tier_for_system(access_assignment.system_id, access_assignment.license_category) or '')
    )
    license_category_value = (
        request.POST.get('license_category', access_assignment.license_category or '').strip()
    )

    if request.method == 'POST':
        access_type = request.POST.get('access_type')
        request_type = request.POST.get('request_type')
        priority = request.POST.get('priority')
        business_justification = request.POST.get('business_justification')
        legitimate_business_need = request.POST.get('legitimate_business_need', '').strip()
        technical_requirements = request.POST.get('technical_requirements')
        requested_access_duration = request.POST.get('requested_access_duration')
        access_start_date = request.POST.get('access_start_date')
        access_end_date = request.POST.get('access_end_date')
        status = request.POST.get('status') or access_assignment.status
        access_username_raw = request.POST.get('access_username')
        system_username = request.POST.get('system_username', '').strip()
        access_url = request.POST.get('access_url')
        granted_access_level = request.POST.get('granted_access_level')
        security_clearance_required = request.POST.get('security_clearance_required')
        data_access_level = request.POST.get('data_access_level')
        risk_assessment_score = request.POST.get('risk_assessment_score')
        review_frequency_days = request.POST.get('review_frequency_days')
        special_instructions = request.POST.get('special_instructions')
        compliance_requirements = request.POST.get('compliance_requirements')
        username_verified_by_id = request.POST.get('username_verified_by')
        username_verified_date_raw = request.POST.get('username_verified_date')
        username_verification_artifact_url = request.POST.get('username_verification_artifact_url', '').strip()
        verification_artifact_file = request.FILES.get('username_verification_artifact')
        clear_verification_artifact = request.POST.get('clear_username_verification_artifact') == 'on'

        # System Owner authorization fields (RHG 4.4)
        system_owner_approved = request.POST.get('system_owner_approved') == 'on'
        system_owner_approval_date_raw = request.POST.get('system_owner_approval_date')
        system_owner_approver_id = request.POST.get('system_owner_approver')
        
        # Generic account fields
        is_generic = request.POST.get('is_generic_account') == 'on'
        generic_remediated = request.POST.get('generic_account_remediated') == 'on'
        remediation_date = request.POST.get('remediation_date')
        remediation_notes = request.POST.get('remediation_notes', '')

        # Administrator-equivalent access metadata (RHG 4.3)
        is_admin_access = (
            request.POST.get('is_admin_access') == 'on'
            or access_type in ['Admin', 'Super Admin']
        )
        has_separate_admin_account = request.POST.get('has_separate_admin_account') == 'on'
        admin_account_username = request.POST.get('admin_account_username', '').strip()
        regular_account_username = request.POST.get('regular_account_username', '').strip()
        is_workstation_login = request.POST.get('is_workstation_login') == 'on'
        has_domain_admin = request.POST.get('has_domain_admin') == 'on'
        admin_password_storage_location = request.POST.get('admin_password_storage_location', '').strip()
        admin_password_stored_date_raw = request.POST.get('admin_password_stored_date')
        manual_license_category = request.POST.get('license_category', '').strip()

        # Resolve subscription tier mapping for subscription-based systems
        tiers_for_system = list(
            SystemSubscriptionTier.objects.select_related('contract').filter(contract__system=access_assignment.system)
        )
        tier_lookup = {str(tier.id): tier for tier in tiers_for_system}
        selected_tier_obj = (
            tier_lookup.get(str(selected_subscription_tier_id)) if selected_subscription_tier_id else None
        )
        if not selected_tier_obj and not manual_license_category and len(tiers_for_system) == 1:
            selected_tier_obj = tiers_for_system[0]
            selected_subscription_tier_id = str(selected_tier_obj.id)
        if selected_tier_obj:
            license_category_value = selected_tier_obj.license_category or selected_tier_obj.name
        elif manual_license_category:
            license_category_value = manual_license_category
        
        # If system_username is empty but access_username exists, use access_username
        # This helps migrate legacy data
        if not system_username and access_username_raw:
            system_username = access_username_raw.strip()
        
        # Convert empty string to None for database consistency
        system_username = system_username if system_username else None
        if access_username_raw is None:
            access_username = access_assignment.access_username
        else:
            access_username = access_username_raw.strip() or None
        
        try:
            # Update fields
            access_assignment.access_type = access_type
            access_assignment.request_type = request_type or access_assignment.request_type
            access_assignment.priority = priority
            access_assignment.business_justification = business_justification
            access_assignment.legitimate_business_need = legitimate_business_need or None
            access_assignment.technical_requirements = technical_requirements
            access_assignment.requested_access_duration = int(requested_access_duration) if requested_access_duration else None
            access_assignment.access_start_date = timezone.datetime.fromisoformat(access_start_date) if access_start_date else None
            access_assignment.access_end_date = timezone.datetime.fromisoformat(access_end_date) if access_end_date else None
            access_assignment.status = status
            access_assignment.updated_by = request.user
            access_assignment.access_username = access_username
            access_assignment.system_username = system_username
            access_assignment.access_url = access_url
            access_assignment.granted_access_level = granted_access_level
            access_assignment.security_clearance_required = security_clearance_required
            access_assignment.data_access_level = data_access_level
            access_assignment.risk_assessment_score = int(risk_assessment_score) if risk_assessment_score else None
            access_assignment.review_frequency_days = int(review_frequency_days) if review_frequency_days else None
            access_assignment.special_instructions = special_instructions
            access_assignment.compliance_requirements = compliance_requirements
            access_assignment.license_category = license_category_value or None

            # System Owner authorization metadata
            access_assignment.system_owner_approved = system_owner_approved
            access_assignment.system_owner_approval_date = _parse_datetime_input(system_owner_approval_date_raw)
            system_owner_approver = None
            if system_owner_approver_id:
                system_owner_approver = CustomUser.objects.filter(id=system_owner_approver_id).first()
            access_assignment.system_owner_approver = system_owner_approver

            # Update admin access metadata
            access_assignment.is_admin_access = is_admin_access
            access_assignment.has_separate_admin_account = has_separate_admin_account
            access_assignment.admin_account_username = admin_account_username or None
            access_assignment.regular_account_username = regular_account_username or None
            access_assignment.is_workstation_login = is_workstation_login
            access_assignment.has_domain_admin = has_domain_admin
            access_assignment.admin_password_storage_location = admin_password_storage_location or None

            if admin_password_stored_date_raw:
                access_assignment.admin_password_stored_date = _parse_datetime_input(admin_password_stored_date_raw)
                if not access_assignment.admin_password_stored_by:
                    access_assignment.admin_password_stored_by = request.user
            
            # Update generic account fields
            access_assignment.is_generic_account = is_generic
            access_assignment.generic_account_remediated = generic_remediated
            if remediation_date:
                access_assignment.remediation_date = timezone.datetime.fromisoformat(remediation_date)
            access_assignment.remediation_notes = remediation_notes
            if generic_remediated and not access_assignment.remediated_by:
                access_assignment.remediated_by = request.user

            # Username verification metadata
            username_verified_by = None
            if username_verified_by_id:
                username_verified_by = CustomUser.objects.filter(id=username_verified_by_id).first()
            if not username_verified_by and username_verified_date_raw:
                username_verified_by = request.user

            access_assignment.username_verified_by = username_verified_by
            access_assignment.username_verified_date = _parse_datetime_input(username_verified_date_raw)
            access_assignment.username_verification_artifact_url = username_verification_artifact_url or None

            if verification_artifact_file:
                if access_assignment.username_verification_artifact:
                    access_assignment.username_verification_artifact.delete(save=False)
                access_assignment.username_verification_artifact = verification_artifact_file
            elif clear_verification_artifact and access_assignment.username_verification_artifact:
                access_assignment.username_verification_artifact.delete(save=False)
                access_assignment.username_verification_artifact = None
            
            # Auto-detect generic accounts
            if system_username:
                access_assignment.mark_as_generic_if_needed()
                if access_assignment.is_generic_account and not is_generic:
                    messages.warning(
                        request,
                        f'Warning: Username "{system_username}" appears to be a generic account. '
                        f'It has been automatically flagged.'
                    )
            
            access_assignment.save()
            
            # Create access history entry
            AccessHistory.objects.create(
                user=access_assignment.user,
                system=access_assignment.system,
                user_system_access=access_assignment,
                action='Modified',
                action_description=f'Access modified by {request.user.full_name}',
                created_by=request.user
            )
            
            messages.success(request, 'Access assignment updated successfully.')
            return redirect('access_management:access_assignment_detail', pk=access_assignment.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating access assignment: {str(e)}')
    
    users_queryset = CustomUser.objects.all().order_by('first_name', 'last_name')
    systems_queryset = System.objects.all().order_by('name')

    context = {
        'access_assignment': access_assignment,
        # dropdown choices
        'access_type_choices': UserSystemAccess.ACCESS_TYPE_CHOICES,
        'request_type_choices': UserSystemAccess.REQUEST_TYPE_CHOICES,
        'priority_choices': UserSystemAccess.PRIORITY_CHOICES,
        'status_choices': UserSystemAccess.STATUS_CHOICES,
        # pre-selected values for template
        'selected_user_id': str(access_assignment.user_id),
        'selected_system_id': str(access_assignment.system_id),
        'selected_access_type': access_assignment.access_type,
        'selected_request_type': access_assignment.request_type,
        'selected_priority': access_assignment.priority,
        # lists
        'users': users_queryset,
        'verifiers': users_queryset,
        'systems': systems_queryset,
        'business_justification_value': request.POST.get('business_justification', access_assignment.business_justification or ''),
        'requested_access_duration_value': request.POST.get('requested_access_duration', access_assignment.requested_access_duration or ''),
        'access_start_date_value': request.POST.get(
            'access_start_date',
            timezone.localtime(access_assignment.access_start_date).strftime('%Y-%m-%dT%H:%M') if access_assignment.access_start_date else ''
        ),
        'access_end_date_value': request.POST.get(
            'access_end_date',
            timezone.localtime(access_assignment.access_end_date).strftime('%Y-%m-%dT%H:%M') if access_assignment.access_end_date else ''
        ),
        # Pre-populate system_username from effective_username if system_username is empty (for legacy data migration)
        'system_username_value': request.POST.get('system_username', access_assignment.effective_username or ''),
        'license_category_value': license_category_value,
        'subscription_tiers': _get_subscription_tier_options(),
        'selected_subscription_tier_id': str(selected_subscription_tier_id),
        'access_url_value': request.POST.get('access_url', access_assignment.access_url or ''),
        'granted_access_level_value': request.POST.get('granted_access_level', access_assignment.granted_access_level or ''),
        'technical_requirements_value': request.POST.get('technical_requirements', access_assignment.technical_requirements or ''),
        'security_clearance_required_value': request.POST.get('security_clearance_required', access_assignment.security_clearance_required or ''),
        'data_access_level_value': request.POST.get('data_access_level', access_assignment.data_access_level or ''),
        'risk_assessment_score_value': request.POST.get('risk_assessment_score', access_assignment.risk_assessment_score or ''),
        'review_frequency_days_value': request.POST.get('review_frequency_days', access_assignment.review_frequency_days or ''),
        'special_instructions_value': request.POST.get('special_instructions', access_assignment.special_instructions or ''),
        'compliance_requirements_value': request.POST.get('compliance_requirements', access_assignment.compliance_requirements or ''),
        # System Owner authorization (4.4)
        'system_owner_approved_value': request.POST.get(
            'system_owner_approved',
            'on' if access_assignment.system_owner_approved else '',
        ),
        'system_owner_approval_date_value': request.POST.get(
            'system_owner_approval_date',
            _format_datetime_for_input(access_assignment.system_owner_approval_date),
        ),
        'system_owner_approver_value': request.POST.get(
            'system_owner_approver',
            str(access_assignment.system_owner_approver_id) if access_assignment.system_owner_approver_id else '',
        ),
        'legitimate_business_need_value': request.POST.get(
            'legitimate_business_need',
            access_assignment.legitimate_business_need or '',
        ),
        # Admin access governance values
        'is_admin_access_value': request.POST.get(
            'is_admin_access',
            'on' if access_assignment.is_admin_access or access_assignment.access_type in ['Admin', 'Super Admin'] else ''
        ),
        'has_separate_admin_account_value': request.POST.get(
            'has_separate_admin_account',
            'on' if access_assignment.has_separate_admin_account else ''
        ),
        'admin_account_username_value': request.POST.get(
            'admin_account_username',
            access_assignment.admin_account_username or ''
        ),
        'regular_account_username_value': request.POST.get(
            'regular_account_username',
            access_assignment.regular_account_username or ''
        ),
        'is_workstation_login_value': request.POST.get(
            'is_workstation_login',
            'on' if access_assignment.is_workstation_login else ''
        ),
        'has_domain_admin_value': request.POST.get(
            'has_domain_admin',
            'on' if access_assignment.has_domain_admin else ''
        ),
        'admin_password_storage_location_value': request.POST.get(
            'admin_password_storage_location',
            access_assignment.admin_password_storage_location or ''
        ),
        'admin_password_stored_date_value': request.POST.get(
            'admin_password_stored_date',
            _format_datetime_for_input(access_assignment.admin_password_stored_date)
        ),
        'username_verified_by_value': request.POST.get(
            'username_verified_by',
            str(access_assignment.username_verified_by_id) if access_assignment.username_verified_by_id else ''
        ),
        'username_verified_date_value': request.POST.get(
            'username_verified_date',
            _format_datetime_for_input(access_assignment.username_verified_date)
        ),
        'username_verification_artifact_url_value': request.POST.get(
            'username_verification_artifact_url',
            access_assignment.username_verification_artifact_url or ''
        ),
    }
    
    return render(request, 'access_management/access_assignment_form.html', context)


@login_required
def access_assignment_delete(request, pk):
    """Delete an access assignment"""
    access_assignment = get_object_or_404(UserSystemAccess, pk=pk)
    
    if request.method == 'POST':
        try:
            # Create access history entry before deletion
            AccessHistory.objects.create(
                user=access_assignment.user,
                system=access_assignment.system,
                action='Revoked',
                action_description=f'Access revoked by {request.user.full_name}',
                created_by=request.user
            )
            
            access_assignment.delete()
            messages.success(request, 'Access assignment deleted successfully.')
            return redirect('access_management:access_assignment_list')
            
        except Exception as e:
            messages.error(request, f'Error deleting access assignment: {str(e)}')
    
    context = {
        'access_assignment': access_assignment,
    }
    
    return render(request, 'access_management/access_assignment_confirm_delete.html', context)


@login_required
def user_access_assignments(request, user_id):
    """View and manage access assignments for a specific user"""
    user = get_object_or_404(CustomUser, pk=user_id)
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    system_filter = request.GET.get('system', '')
    next_url = request.GET.get('next', '')
    
    # Base queryset - optimized with select_related for efficiency
    queryset = UserSystemAccess.objects.select_related(
        'system', 'approved_by', 'requested_by', 'updated_by'
    ).filter(user=user)
    
    # Apply filters
    if status_filter and status_filter.lower() != 'all':
        queryset = queryset.filter(status__iexact=status_filter)
    if system_filter:
        queryset = queryset.filter(system_id=system_filter)

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_assignments')
        action = request.POST.get('bulk_action')

        if not selected_ids:
            messages.error(request, 'Select at least one assignment to perform bulk actions.')
            return redirect(request.get_full_path())

        assignments = UserSystemAccess.objects.filter(pk__in=selected_ids, user=user)

        if action == 'bulk_delete':
            deleted = 0
            for assignment in assignments:
                AccessHistory.objects.create(
                    user=assignment.user,
                    system=assignment.system,
                    action='Revoked',
                    action_description=f'Access revoked in bulk by {request.user.full_name}',
                    created_by=request.user
                )
                assignment.delete()
                deleted += 1

            messages.success(request, f'Successfully deleted {deleted} access assignment{"s" if deleted != 1 else ""}.')
            return redirect(request.path_info)

        if action == 'bulk_update':
            new_status = request.POST.get('bulk_status', '')
            new_priority = request.POST.get('bulk_priority', '')
            new_access_type = request.POST.get('bulk_access_type', '')

            if not any([new_status, new_priority, new_access_type]):
                messages.error(request, 'Choose at least one field to update when performing a bulk edit.')
                return redirect(request.get_full_path())

            updated = 0
            for assignment in assignments:
                changes = []
                if new_status:
                    assignment.status = new_status
                    changes.append(f"status → {new_status}")
                if new_priority:
                    assignment.priority = new_priority
                    changes.append(f"priority → {new_priority}")
                if new_access_type:
                    assignment.access_type = new_access_type
                    changes.append(f"access type → {new_access_type}")

                if changes:
                    assignment.updated_by = request.user
                    assignment.save()
                    AccessHistory.objects.create(
                        user=assignment.user,
                        system=assignment.system,
                        user_system_access=assignment,
                        action='Modified',
                        action_description=f"Bulk update by {request.user.full_name}: {', '.join(changes)}",
                        created_by=request.user
                    )
                    updated += 1

            messages.success(request, f'Successfully updated {updated} access assignment{"s" if updated != 1 else ""}.')
            return redirect(request.get_full_path())

        messages.error(request, 'Unsupported bulk action.')
        return redirect(request.get_full_path())
    
    # Pagination - order by system for accordion grouping in template
    queryset_ordered = queryset.order_by('system__name', '-request_date', 'pk')
    paginator = Paginator(queryset_ordered, 25)
    page_number = request.GET.get('page')
    access_assignments = paginator.get_page(page_number)
    
    # Summary metrics
    total_assignments = queryset.count()
    active_assignments = queryset.filter(status__in=['Active', 'Approved']).count()
    pending_assignments = queryset.filter(status='Pending').count()
    expired_assignments = queryset.filter(status='Expired').count()
    unique_systems = queryset.values('system_id').distinct().count()
    
    # Get user's systems for filter
    user_systems = System.objects.filter(
        user_accesses__user=user
    ).distinct().order_by('name')
    
    # Fallback next_url to access_assignment_list if not provided
    if not next_url:
        next_url = request.META.get('HTTP_REFERER', '')
        if 'user/' in next_url and f'/users/{user_id}/' in next_url:
            # Avoid redirect loop back to same page
            next_url = ''
    
    context = {
        'user': user,
        'access_assignments': access_assignments,
        'status_choices': UserSystemAccess.STATUS_CHOICES,
        'priority_choices': UserSystemAccess.PRIORITY_CHOICES,
        'access_type_choices': UserSystemAccess.ACCESS_TYPE_CHOICES,
        'systems': user_systems,
        'filters': {
            'status': status_filter,
            'system': system_filter,
        },
        'next_url': next_url,
        'total_assignments': total_assignments,
        'active_assignments': active_assignments,
        'pending_assignments': pending_assignments,
        'expired_assignments': expired_assignments,
        'unique_systems': unique_systems,
        'is_paginated': access_assignments.has_other_pages(),
        'page_obj': access_assignments,
    }
    
    return render(request, 'access_management/user_access_assignments.html', context)


@login_required
def system_access_assignments(request, system_id):
    """View and manage access assignments for a specific system"""
    system = get_object_or_404(System, pk=system_id)
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    access_type_filter = request.GET.get('access_type', '')
    
    # Base queryset
    queryset = UserSystemAccess.objects.select_related('user').filter(system=system)
    
    # Apply filters
    if status_filter and status_filter.lower() != 'all':
        queryset = queryset.filter(status__iexact=status_filter)
    if access_type_filter:
        queryset = queryset.filter(access_type=access_type_filter)

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_assignments')
        action = request.POST.get('bulk_action')

        if not selected_ids:
            messages.error(request, 'Select at least one assignment to perform bulk actions.')
            return redirect(request.get_full_path())

        assignments = UserSystemAccess.objects.filter(pk__in=selected_ids, system=system)

        if action == 'bulk_delete':
            deleted = 0
            for assignment in assignments:
                AccessHistory.objects.create(
                    user=assignment.user,
                    system=assignment.system,
                    action='Revoked',
                    action_description=f'Access revoked in bulk by {request.user.full_name}',
                    created_by=request.user
                )
                assignment.delete()
                deleted += 1

            messages.success(request, f'Successfully deleted {deleted} access assignment{"s" if deleted != 1 else ""}.')
            return redirect(request.path_info)

        if action == 'bulk_update':
            new_status = request.POST.get('bulk_status', '')
            new_priority = request.POST.get('bulk_priority', '')
            new_access_type = request.POST.get('bulk_access_type', '')

            if not any([new_status, new_priority, new_access_type]):
                messages.error(request, 'Choose at least one field to update when performing a bulk edit.')
                return redirect(request.get_full_path())

            updated = 0
            for assignment in assignments:
                changes = []
                if new_status:
                    assignment.status = new_status
                    changes.append(f"status → {new_status}")
                if new_priority:
                    assignment.priority = new_priority
                    changes.append(f"priority → {new_priority}")
                if new_access_type:
                    assignment.access_type = new_access_type
                    changes.append(f"access type → {new_access_type}")

                if changes:
                    assignment.updated_by = request.user
                    assignment.save()
                    AccessHistory.objects.create(
                        user=assignment.user,
                        system=assignment.system,
                        user_system_access=assignment,
                        action='Modified',
                        action_description=f"Bulk update by {request.user.full_name}: {', '.join(changes)}",
                        created_by=request.user
                    )
                    updated += 1

            messages.success(request, f'Successfully updated {updated} access assignment{"s" if updated != 1 else ""}.')
            return redirect(request.get_full_path())

        messages.error(request, 'Unsupported bulk action.')
        return redirect(request.get_full_path())
    
    # Pagination
    paginator = Paginator(queryset.order_by('-created_at'), 25)
    page_number = request.GET.get('page')
    access_assignments = paginator.get_page(page_number)
    
    # Summary metrics
    total_assignments = queryset.count()
    active_assignments = queryset.filter(status='Active').count()
    pending_assignments = queryset.filter(status='Pending').count()
    unique_users = queryset.values('user_id').distinct().count()
    
    context = {
        'system': system,
        'access_assignments': access_assignments,
        'status_choices': UserSystemAccess.STATUS_CHOICES,
        'priority_choices': UserSystemAccess.PRIORITY_CHOICES,
        'access_type_choices': UserSystemAccess.ACCESS_TYPE_CHOICES,
        'filters': {
            'status': status_filter,
            'access_type': access_type_filter,
        },
        'total_assignments': total_assignments,
        'active_assignments': active_assignments,
        'pending_assignments': pending_assignments,
        'unique_users': unique_users,
        'access_levels': {
            (item['granted_access_level'] or 'Unspecified'): item['count']
            for item in queryset.values('granted_access_level').annotate(count=Count('id')).order_by('granted_access_level')
        },
        'is_paginated': access_assignments.has_other_pages(),
        'page_obj': access_assignments,
    }
    
    return render(request, 'access_management/system_access_assignments.html', context)


@login_required
def approve_access_assignment(request, pk):
    """Approve an access assignment"""
    access_assignment = get_object_or_404(UserSystemAccess, pk=pk)
    
    if request.method == 'POST':
        comments = request.POST.get('approval_comments', '')
        
        try:
            access_assignment.approve_access(request.user, comments)
            
            # Create access history entry
            AccessHistory.objects.create(
                user=access_assignment.user,
                system=access_assignment.system,
                user_system_access=access_assignment,
                action='Approved',
                action_description=f'Access approved by {request.user.full_name}',
                created_by=request.user
            )
            
            messages.success(request, f'Access assignment approved for {access_assignment.user.full_name}.')
            
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error approving access assignment: {str(e)}')
    
    return redirect('access_management:access_assignment_detail', pk=pk)


@login_required
def reject_access_assignment(request, pk):
    """Reject an access assignment"""
    access_assignment = get_object_or_404(UserSystemAccess, pk=pk)
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        
        if not rejection_reason:
            messages.error(request, 'Rejection reason is required.')
            return redirect('access_management:access_assignment_detail', pk=pk)
        
        try:
            access_assignment.reject_access(request.user, rejection_reason)
            
            # Create access history entry
            AccessHistory.objects.create(
                user=access_assignment.user,
                system=access_assignment.system,
                user_system_access=access_assignment,
                action='Rejected',
                action_description=f'Access rejected by {request.user.full_name}',
                created_by=request.user
            )
            
            messages.success(request, f'Access assignment rejected for {access_assignment.user.full_name}.')
            
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error rejecting access assignment: {str(e)}')
    
    return redirect('access_management:access_assignment_detail', pk=pk)


@login_required
def access_history_list(request):
    """List all access history events"""
    
    # Get filter parameters
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')
    system_filter = request.GET.get('system', '')
    success_filter = request.GET.get('success', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    queryset = AccessHistory.objects.select_related('user', 'system', 'user_system_access').all()
    
    # Apply filters
    if action_filter:
        queryset = queryset.filter(action=action_filter)
    if user_filter:
        queryset = queryset.filter(user_id=user_filter)
    if system_filter:
        queryset = queryset.filter(system_id=system_filter)
    if success_filter:
        queryset = queryset.filter(success=success_filter.lower() == 'true')
    if search_query:
        queryset = queryset.filter(
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(system__name__icontains=search_query) |
            Q(action_description__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    access_history = paginator.get_page(page_number)
    
    # Get filter options
    systems = System.objects.all().order_by('name')
    users = CustomUser.objects.all().order_by('first_name', 'last_name')
    
    context = {
        'access_history': access_history,
        'action_choices': AccessHistory.ACTION_CHOICES,
        'systems': systems,
        'users': users,
        'filters': {
            'action': action_filter,
            'user': user_filter,
            'system': system_filter,
            'success': success_filter,
            'search': search_query,
        }
    }
    
    return render(request, 'access_management/access_history_list.html', context)


@login_required
def user_access_history(request, user_id):
    """Display access history for a specific user."""
    user = get_object_or_404(CustomUser, pk=user_id)
    queryset = AccessHistory.objects.filter(user=user).select_related('system').order_by('-timestamp')
    
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    access_history = paginator.get_page(page_number)
    
    context = {
        'access_history': access_history,
        'user': user,
        'title': f'Access History for {user.get_full_name()}',
    }
    
    return render(request, 'access_management/access_history_list.html', context)


@login_required
def system_access_history(request, system_id):
    """Display access history for a specific system."""
    system = get_object_or_404(System, pk=system_id)
    queryset = AccessHistory.objects.filter(system=system).select_related('user').order_by('-timestamp')
    
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    access_history = paginator.get_page(page_number)
    
    context = {
        'access_history': access_history,
        'system': system,
        'title': f'Access History for {system.name}',
    }
    
    return render(request, 'access_management/access_history_list.html', context)


@login_required
def assignment_access_history(request, assignment_id):
    """Display access history for a specific assignment."""
    assignment = get_object_or_404(UserSystemAccess, pk=assignment_id)
    queryset = AccessHistory.objects.filter(user_system_access=assignment).select_related('user', 'system').order_by('-timestamp')
    
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    access_history = paginator.get_page(page_number)
    
    context = {
        'access_history': access_history,
        'assignment': assignment,
        'title': f'Access History for Assignment {assignment.id}',
    }
    
    return render(request, 'access_management/access_history_list.html', context)


@login_required
def generic_accounts_report(request):
    """Report of all generic accounts across external systems"""
    # Get filter parameters
    system_id = request.GET.get('system')
    show_remediated = request.GET.get('show_remediated', 'false') == 'true'
    search = request.GET.get('search', '').strip()
    
    # Start with all generic accounts
    queryset = UserSystemAccess.objects.filter(is_generic_account=True).select_related('user', 'system')
    
    # Filter by system
    if system_id:
        queryset = queryset.filter(system_id=system_id)
    
    # Filter by remediation status
    if not show_remediated:
        queryset = queryset.filter(generic_account_remediated=False)
    
    # Search filter - search in both system_username and access_username (for legacy data)
    if search:
        queryset = queryset.filter(
            Q(system_username__icontains=search) |
            Q(access_username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(system__name__icontains=search)
        )
    
    # Order by system, then username (use system_username first, fallback to access_username)
    # Note: We can't order by effective_username directly, so we order by system_username
    # Records with only access_username will appear after those with system_username
    queryset = queryset.order_by('system__name', 'system_username', 'access_username')
    
    # Pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_generic = UserSystemAccess.objects.filter(is_generic_account=True).count()
    unremediated = UserSystemAccess.objects.filter(
        is_generic_account=True,
        generic_account_remediated=False
    ).count()
    remediated = total_generic - unremediated
    
    # Group by system
    by_system = queryset.values('system__name', 'system__code').annotate(
        count=Count('id')
    ).order_by('system__name')
    
    context = {
        'page_obj': page_obj,
        'generic_accounts': page_obj,
        'systems': System.objects.all().order_by('name'),
        'selected_system_id': system_id,
        'show_remediated': show_remediated,
        'search': search,
        'total_generic': total_generic,
        'unremediated': unremediated,
        'remediated': remediated,
        'by_system': by_system,
    }
    
    return render(request, 'access_management/generic_accounts_report.html', context)


@login_required
def quarterly_access_review_dashboard(request):
    """
    Dashboard for RHG 4.5 quarterly review documentation & permission change logging.
    """
    default_quarter = get_current_quarter_label()
    selected_quarter = request.GET.get('quarter') or default_quarter
    system_filter = request.GET.get('system', '').strip()
    match_filter = request.GET.get('match', 'all')
    owner_filter = request.GET.get('owner', 'all')
    status_filter = request.GET.get('status', 'all')
    export_format = request.GET.get('export')

    reviews_qs = QuarterlyAccessReview.objects.select_related(
        'reviewed_user',
        'reviewed_by',
        'system',
        'system_owner',
        'user_system_access',
    )

    include_all_quarters = selected_quarter == 'all'
    if not include_all_quarters:
        reviews_qs = reviews_qs.filter(review_quarter=selected_quarter)

    if system_filter:
        reviews_qs = reviews_qs.filter(system_id=system_filter)

    if match_filter == 'match':
        reviews_qs = reviews_qs.filter(matches_approved=True)
    elif match_filter == 'mismatch':
        reviews_qs = reviews_qs.filter(matches_approved=False)

    if owner_filter == 'confirmed':
        reviews_qs = reviews_qs.filter(system_owner_confirmed=True)
    elif owner_filter == 'unconfirmed':
        reviews_qs = reviews_qs.filter(system_owner_confirmed=False)

    if status_filter == 'completed':
        reviews_qs = reviews_qs.filter(review_completed=True)
    elif status_filter == 'pending':
        reviews_qs = reviews_qs.filter(review_completed=False)

    if export_format == 'csv':
        return export_quarterly_reviews_to_csv(reviews_qs)

    metrics = {
        'total_reviews': reviews_qs.count(),
        'completed_reviews': reviews_qs.filter(review_completed=True).count(),
        'owner_confirmed': reviews_qs.filter(system_owner_confirmed=True).count(),
        'mismatches': reviews_qs.filter(matches_approved=False).count(),
    }

    annual_progress = _annual_review_progress()

    paginator = Paginator(reviews_qs.order_by('-review_date'), 25)
    reviews_page = paginator.get_page(request.GET.get('page'))

    permission_changes_qs = PermissionChangeDocumentation.objects.select_related(
        'user_system_access__user',
        'user_system_access__system',
        'approval_reference',
        'documented_by',
    )
    if system_filter:
        permission_changes_qs = permission_changes_qs.filter(user_system_access__system_id=system_filter)
    date_range = _quarter_date_range(selected_quarter)
    if date_range[0] and date_range[1]:
        permission_changes_qs = permission_changes_qs.filter(
            changed_in_external_system_date__range=date_range
        )
    recent_permission_changes = permission_changes_qs.order_by('-changed_in_external_system_date')[:10]

    now = timezone.now()
    upcoming_window = now + timedelta(days=30)
    due_assignments = UserSystemAccess.objects.select_related('user', 'system').filter(
        status__in=['Active', 'Approved'],
        next_review_date__isnull=False,
        next_review_date__lte=upcoming_window,
    ).order_by('next_review_date')
    overdue_assignment_count = due_assignments.filter(next_review_date__lt=now).count()
    due_assignment_count = due_assignments.count()
    pending_assignments = due_assignments[:10]

    review_form_initial = {
        'review_quarter': selected_quarter if selected_quarter != 'all' else default_quarter,
        'reviewed_by': request.user.pk,
        'review_date': timezone.now().strftime('%Y-%m-%dT%H:%M'),
    }
    review_form = QuarterlyAccessReviewForm(initial=review_form_initial)
    permission_change_form = PermissionChangeDocumentationForm(initial={'documented_by': request.user.pk})

    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        if form_type == 'quarterly_review':
            review_form = QuarterlyAccessReviewForm(request.POST)
            if review_form.is_valid():
                review_instance = review_form.save(commit=False)
                if not review_instance.reviewed_by:
                    review_instance.reviewed_by = request.user
                if review_instance.system_owner_confirmed and not review_instance.system_owner_confirmed_date:
                    review_instance.system_owner_confirmed_date = timezone.now()
                review_instance.save()
                messages.success(request, 'Quarterly access review logged successfully.')
                return redirect(request.get_full_path())
            messages.error(request, 'Please correct the errors in the quarterly review form.')
        elif form_type == 'permission_change':
            permission_change_form = PermissionChangeDocumentationForm(request.POST)
            if permission_change_form.is_valid():
                permission_change = permission_change_form.save(commit=False)
                if not permission_change.documented_by:
                    permission_change.documented_by = request.user
                permission_change.save()
                messages.success(request, 'Permission change documentation saved.')
                return redirect(request.get_full_path())
            messages.error(request, 'Please correct the errors in the permission change form.')

    quarter_options = ['all'] + _build_quarter_options(8)
    systems = System.objects.filter(is_active=True).order_by('name')

    filters = {
        'quarter': selected_quarter,
        'system': system_filter,
        'match': match_filter,
        'owner': owner_filter,
        'status': status_filter,
    }

    context = {
        'reviews_page': reviews_page,
        'metrics': metrics,
        'quarter_options': quarter_options,
        'systems': systems,
        'filters': filters,
        'review_form': review_form,
        'permission_change_form': permission_change_form,
        'recent_permission_changes': recent_permission_changes,
        'pending_assignments': pending_assignments,
        'overdue_assignment_count': overdue_assignment_count,
        'due_assignment_count': due_assignment_count,
        'annual_progress': annual_progress,
    }
    return render(request, 'access_management/quarterly_access_reviews.html', context)


@login_required
def quarterly_access_review_bulk(request):
    """Bulk-generate quarterly review records for multiple users/systems."""
    initial = {
        'review_quarter': get_current_quarter_label(),
        'reviewed_by': request.user.pk,
        'review_date': timezone.now().strftime('%Y-%m-%dT%H:%M'),
    }
    form = BulkQuarterlyReviewForm(initial=initial)
    preview_assignments = []
    selection_summary = {}
    created_reviews = []
    skipped_assignments = []

    if request.method == 'POST':
        form = BulkQuarterlyReviewForm(request.POST)
        if form.is_valid():
            cleaned = form.cleaned_data
            system = cleaned['system']
            review_quarter = cleaned['review_quarter']
            users_qty = cleaned['users_qty']
            preview_assignments = _select_assignments_for_bulk(system, users_qty, review_quarter)
            selection_summary = {
                'system': system,
                'requested': users_qty,
                'generated': len(preview_assignments),
                'review_quarter': review_quarter,
            }

            if not preview_assignments:
                messages.warning(
                    request,
                    'No eligible assignments found for the selected quarter/system. '
                    'Try increasing the number of users or choosing a different system.',
                )
            elif 'generate' in request.POST:
                review_date = cleaned['review_date']
                reviewed_by = cleaned['reviewed_by']
                matches_approved = cleaned['matches_approved']
                review_completed = cleaned['review_completed']
                discrepancies = cleaned['discrepancies']
                for assignment in preview_assignments:
                    if not assignment.user or not assignment.system:
                        skipped_assignments.append({
                            'label': assignment.user.full_name if assignment.user else 'Unknown User',
                            'reason': 'Missing user/system reference',
                        })
                        continue
                    exists = QuarterlyAccessReview.objects.filter(
                        review_quarter=review_quarter,
                        reviewed_user=assignment.user,
                        system=assignment.system,
                    ).exists()
                    if exists:
                        skipped_assignments.append({
                            'label': assignment.user.full_name if assignment.user else 'Unknown User',
                            'reason': 'Already reviewed this quarter',
                        })
                        continue

                    review = QuarterlyAccessReview(
                        review_quarter=review_quarter,
                        reviewed_user=assignment.user,
                        system=assignment.system,
                        user_system_access=assignment,
                        reviewed_by=reviewed_by,
                        review_date=review_date,
                        approved_permissions=_default_permission_label(assignment),
                        actual_permissions_in_external_system=_default_permission_label(assignment),
                        matches_approved=matches_approved,
                        discrepancies=discrepancies,
                        system_owner=assignment.system.system_owner if assignment.system else None,
                        system_owner_confirmed=False,
                        review_completed=review_completed,
                    )
                    review.save()
                    created_reviews.append(review)
                    _update_assignment_review_schedule(assignment, review_date)

                if created_reviews:
                    messages.success(
                        request,
                        f"{len(created_reviews)} quarterly review record{'s' if len(created_reviews) != 1 else ''} "
                        f"generated for {system.name}.",
                    )
                if skipped_assignments:
                    messages.warning(
                        request,
                        f"{len(skipped_assignments)} assignment{'s' if len(skipped_assignments) != 1 else ''} "
                        "skipped because a quarterly review already exists or data was incomplete.",
                    )
            else:
                messages.info(
                    request,
                    f"Previewing {len(preview_assignments)} assignment{'s' if len(preview_assignments) != 1 else ''} "
                    f"for {system.name}. Click 'Generate Reviews' to create records.",
                )

    annual_progress = _annual_review_progress()
    available_systems = System.objects.filter(is_active=True).order_by('name')

    context = {
        'form': form,
        'preview_assignments': preview_assignments,
        'selection_summary': selection_summary,
        'created_reviews': created_reviews,
        'skipped_assignments': skipped_assignments,
        'annual_progress': annual_progress,
        'systems': available_systems,
    }
    return render(request, 'access_management/bulk_quarterly_reviews.html', context)


@login_required
def access_approval_compliance(request):
    """
    RHG 4.6 dashboard aggregating quarterly active-user reviews,
    monthly obsolete-account reviews, and access removal documentation.
    """
    default_quarter = get_current_quarter_label()
    now = timezone.now()
    default_month = now.strftime("%Y-%m")

    quarterly_form = QuarterlyActiveUserReviewForm(initial={
        'review_quarter': default_quarter,
        'review_date': now.strftime("%Y-%m-%dT%H:%M"),
        'reviewed_by': request.user.pk,
    })
    monthly_form = MonthlyObsoleteAccountReviewForm(initial={
        'review_month': default_month,
        'review_date': now.strftime("%Y-%m-%dT%H:%M"),
        'reviewed_by': request.user.pk,
    })
    removal_form = AccessRemovalDocumentationForm(initial={
        'removed_by': request.user.pk,
    })

    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        if form_type == 'quarterly_active':
            quarterly_form = QuarterlyActiveUserReviewForm(request.POST)
            if quarterly_form.is_valid():
                quarterly_form.save()
                messages.success(request, "Quarterly active-user review documented.")
                return redirect(request.get_full_path())
            messages.error(request, "Please fix validation errors in the quarterly review form.")
        elif form_type == 'monthly_obsolete':
            monthly_form = MonthlyObsoleteAccountReviewForm(request.POST)
            if monthly_form.is_valid():
                monthly_form.save()
                messages.success(request, "Monthly obsolete account review saved.")
                return redirect(request.get_full_path())
            messages.error(request, "Please fix validation errors in the monthly review form.")
        elif form_type == 'access_removal':
            removal_form = AccessRemovalDocumentationForm(request.POST)
            if removal_form.is_valid():
                instance = removal_form.save(commit=False)
                if instance.verified_removal and not instance.verified_date:
                    instance.verified_date = timezone.now()
                instance.save()
                messages.success(request, "Access removal documentation recorded.")
                return redirect(request.get_full_path())
            messages.error(request, "Please fix validation errors in the removal documentation form.")

    quarterly_reviews = QuarterlyActiveUserReview.objects.select_related('system', 'reviewed_by')[:20]
    monthly_reviews = MonthlyObsoleteAccountReview.objects.select_related('reviewed_by')[:12]
    removal_logs = AccessRemovalDocumentation.objects.select_related(
        'user_system_access__user', 'user_system_access__system', 'removed_by', 'verified_by'
    )[:20]

    unapproved_access_qs = get_unapproved_access_records()
    unapproved_access = unapproved_access_qs[:20]

    obsolete_accounts = identify_obsolete_accounts()
    obsolete_summary = {
        'terminated_users': obsolete_accounts['terminated_users'].count(),
        'inactive_users': obsolete_accounts['inactive_users'].count(),
        'expired_assignments': obsolete_accounts['expired_assignments'].count(),
        'stale_reviews': obsolete_accounts['stale_reviews'].count(),
    }

    total_active_assignments = UserSystemAccess.objects.filter(status__in=['Active', 'Approved']).count()
    removal_pending = AccessRemovalDocumentation.objects.filter(verified_removal=False).count()

    metrics = {
        'total_active_assignments': total_active_assignments,
        'unapproved_access': unapproved_access_qs.count(),
        'quarterly_reviews': QuarterlyActiveUserReview.objects.count(),
        'monthly_reviews': MonthlyObsoleteAccountReview.objects.count(),
        'pending_removals': removal_pending,
    }

    systems = System.objects.filter(name__in=[
        "Active Directory",
        "EMMA CRS",
        "PMS",
        "POS",
        "Doorlock systems",
        "PeopleSearch",
        "Hotelkit",
        "PMI",
        "VPN",
        "OTA's",
        "Credit card portal",
        "Google My Business",
        "Banking software",
        "Accounting software",
    ]).order_by('name')

    context = {
        'metrics': metrics,
        'quarterly_form': quarterly_form,
        'monthly_form': monthly_form,
        'removal_form': removal_form,
        'quarterly_reviews': quarterly_reviews,
        'monthly_reviews': monthly_reviews,
        'removal_logs': removal_logs,
        'unapproved_access': unapproved_access,
        'obsolete_accounts': obsolete_accounts,
        'obsolete_summary': obsolete_summary,
        'systems': systems,
    }
    return render(request, 'access_management/access_approval_compliance.html', context)


@login_required
def policy_drift_monitoring(request):
    """
    Highlight access records that indicate potential policy drift:
    - Accounts missing external usernames
    - Accounts with stale or overdue reviews
    - Usernames reused by multiple employees in the same system
    """
    default_threshold_days = 90

    system_id_param = request.GET.get('system')
    department_id_param = request.GET.get('department')
    status_scope = request.GET.get('status_scope', 'active')
    threshold_param = request.GET.get('stale_threshold')

    def _coerce_int(value):
        if not value:
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    system_id = _coerce_int(system_id_param)
    department_id = _coerce_int(department_id_param)

    try:
        stale_threshold_days = int(threshold_param) if threshold_param else default_threshold_days
        if stale_threshold_days <= 0:
            raise ValueError
    except (TypeError, ValueError):
        stale_threshold_days = default_threshold_days

    snapshot = build_policy_drift_snapshot(
        system_id=system_id,
        department_id=department_id,
        status_scope=status_scope,
        stale_threshold_days=stale_threshold_days,
    )

    def _get_bool(name, default=True):
        values = request.GET.getlist(name)
        if not values:
            return default
        return any(value.lower() == 'true' for value in values)

    show_missing = _get_bool('show_missing', True)
    show_stale = _get_bool('show_stale', True)
    show_overlapping = _get_bool('show_overlapping', True)

    export_format = request.GET.get('export')
    rows = list(generate_policy_drift_rows(snapshot))

    def filter_rows(row_list):
        issue_type = row_list.get('issue_type') if isinstance(row_list, dict) else None
        if issue_type == 'Missing Username' and not show_missing:
            return False
        if issue_type == 'Stale Review' and not show_stale:
            return False
        if issue_type == 'Overlapping Username' and not show_overlapping:
            return False
        return True

    filtered_rows = [row for row in rows if filter_rows(row)]

    if export_format in {'csv', 'pdf'}:
        if export_format == 'csv':
            return export_policy_drift_rows_to_csv(filtered_rows)
        return export_policy_drift_rows_to_pdf(filtered_rows, snapshot['issue_summary'], stale_threshold_days)
    if export_format == 'xlsx':
        return export_policy_drift_rows_to_excel(filtered_rows)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    query_params.pop('export', None)
    base_query = query_params.urlencode()
    def _build_export_link(fmt):
        return f"?{base_query}&export={fmt}" if base_query else f"?export={fmt}"

    csv_link = _build_export_link('csv')
    pdf_link = _build_export_link('pdf')
    xlsx_link = _build_export_link('xlsx')

    context = {
        'issue_summary': snapshot['issue_summary'],
        'missing_usernames': snapshot['missing_usernames_qs'].select_related('user', 'system')[:50] if show_missing else [],
        'stale_reviews': snapshot['stale_reviews_qs'].select_related('user', 'system')[:50] if show_stale else [],
        'overlapping_groups': snapshot['overlapping_groups'].values() if show_overlapping else [],
        'system_issue_counts': snapshot['system_issue_counts'],
        'systems': System.objects.filter(is_active=True).order_by('name'),
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'filters': {
            'system': system_id_param or '',
            'department': department_id_param or '',
            'status_scope': status_scope,
            'stale_threshold': stale_threshold_days,
            'show_missing': show_missing,
            'show_stale': show_stale,
            'show_overlapping': show_overlapping,
        },
        'threshold_options': [30, 60, 90, 120, 180],
        'now': snapshot['now'],
        'stale_reference': snapshot['stale_reference'],
        'export_links': {
            'csv': csv_link,
            'pdf': pdf_link,
            'xlsx': xlsx_link,
        },
    }

    return render(request, 'access_management/policy_drift_monitoring.html', context)


@login_required
def admin_accounts_report(request):
    """
    Administrator-equivalent access compliance report (RHG 4.3).
    Highlights:
    - Admin access limited to IT administrators
    - Separate admin account usage
    - Workstation/domain-admin combinations
    - Admin password storage documentation
    """
    system_id = request.GET.get("system", "").strip()
    department_id = request.GET.get("department", "").strip()
    issue = request.GET.get("issue", "").strip()
    search = request.GET.get("search", "").strip()

    # Base queryset: where is_admin_access is flagged or access_type is Admin/Super Admin
    queryset = UserSystemAccess.objects.select_related("user", "system").filter(
        Q(is_admin_access=True)
        | Q(access_type__in=["Admin", "Super Admin"])
    )

    if system_id:
        queryset = queryset.filter(system_id=system_id)
    if department_id:
        queryset = queryset.filter(user__department_id=department_id)
    if search:
        queryset = queryset.filter(
            Q(user__username__icontains=search)
            | Q(user__first_name__icontains=search)
            | Q(user__last_name__icontains=search)
            | Q(system__name__icontains=search)
            | Q(system_username__icontains=search)
            | Q(admin_account_username__icontains=search)
        )

    # Issue-specific focus
    if issue == "non_it":
        queryset = queryset.filter(
            Q(user__is_it_administrator=False) | Q(user__is_it_administrator__isnull=True)
        )
    elif issue == "no_separate":
        queryset = queryset.filter(
            Q(has_separate_admin_account=False)
            | Q(admin_account_username__isnull=True)
            | Q(admin_account_username__exact="")
        )
    elif issue == "workstation_domain":
        queryset = queryset.filter(
            is_workstation_login=True,
            has_domain_admin=True,
        )
    elif issue == "no_storage":
        queryset = queryset.filter(
            Q(admin_password_storage_location__isnull=True)
            | Q(admin_password_storage_location__exact="")
        )

    queryset = queryset.order_by("user__first_name", "user__last_name", "system__name")

    # Metrics for summary cards (computed on unpaginated queryset within current filter scope)
    base_metrics_qs = queryset
    metrics = {
        "total_admin": base_metrics_qs.count(),
        "non_it_admins": base_metrics_qs.filter(
            Q(user__is_it_administrator=False) | Q(user__is_it_administrator__isnull=True)
        ).count(),
        "no_separate_admin": base_metrics_qs.filter(
            Q(has_separate_admin_account=False)
            | Q(admin_account_username__isnull=True)
            | Q(admin_account_username__exact="")
        ).count(),
        "workstation_domain_admin": base_metrics_qs.filter(
            is_workstation_login=True,
            has_domain_admin=True,
        ).count(),
    }

    export_format = request.GET.get("export")
    if export_format == "xlsx":
        return export_admin_accounts_to_excel(queryset)
    if export_format == "csv":
        return export_admin_accounts_to_csv(queryset)

    paginator = Paginator(queryset, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Build pagination links preserving filters
    query_params = request.GET.copy()
    query_params.pop("page", None)
    base_query = query_params.urlencode()

    def _build_page_link(page_num):
        if not base_query:
            return f"?page={page_num}"
        return f"?{base_query}&page={page_num}"

    pagination = {
        "first": _build_page_link(1),
        "last": _build_page_link(paginator.num_pages) if paginator.num_pages else "",
        "previous": _build_page_link(page_obj.previous_page_number()) if page_obj.has_previous() else "",
        "next": _build_page_link(page_obj.next_page_number()) if page_obj.has_next() else "",
    }

    # Export links
    def _build_export_link(fmt):
        if base_query:
            return f"?{base_query}&export={fmt}"
        return f"?export={fmt}"

    export_links = {
        "xlsx": _build_export_link("xlsx"),
        "csv": _build_export_link("csv"),
    }

    context = {
        "page_obj": page_obj,
        "pagination": pagination,
        "systems": System.objects.filter(is_active=True).order_by("name"),
        "departments": Department.objects.filter(is_active=True).order_by("name"),
        "metrics": metrics,
        "filters": {
            "system": system_id,
            "department": department_id,
            "issue": issue,
            "search": search,
        },
        "export_links": export_links,
    }

    return render(request, "access_management/admin_accounts_report.html", context)


@login_required
def mark_generic_account_remediated(request, pk):
    """Mark a generic account as remediated"""
    access_assignment = get_object_or_404(UserSystemAccess, pk=pk)
    
    if not access_assignment.is_generic_account:
        messages.error(request, 'This account is not marked as generic.')
        return redirect('access_management:generic_accounts_report')
    
    if request.method == 'POST':
        remediation_notes = request.POST.get('remediation_notes', '')
        remediation_date = request.POST.get('remediation_date')
        
        access_assignment.generic_account_remediated = True
        access_assignment.remediated_by = request.user
        access_assignment.remediation_notes = remediation_notes
        if remediation_date:
            try:
                access_assignment.remediation_date = timezone.datetime.fromisoformat(remediation_date.replace('Z', '+00:00'))
            except:
                access_assignment.remediation_date = timezone.now()
        else:
            access_assignment.remediation_date = timezone.now()
        access_assignment.save()
        
        # Create access history entry
        AccessHistory.objects.create(
            user=access_assignment.user,
            system=access_assignment.system,
            user_system_access=access_assignment,
            action='Modified',
            action_description=f'Generic account "{access_assignment.system_username}" marked as remediated by {request.user.full_name}',
            created_by=request.user
        )
        
        messages.success(
            request,
            f'Generic account "{access_assignment.system_username}" has been marked as remediated.'
        )
        return redirect('access_management:generic_accounts_report')
    
    context = {
        'access_assignment': access_assignment,
    }
    
    return render(request, 'access_management/mark_remediated.html', context)


@login_required
def cross_system_account_mapping(request):
    """Cross-system account mapping showing all employees and their usernames across all systems"""
    # Get filter parameters
    user_id = request.GET.get('user')
    system_id = request.GET.get('system')
    department_id = request.GET.get('department')
    search = request.GET.get('search', '').strip()
    show_only_with_access = request.GET.get('show_only_with_access', 'false') == 'true'
    
    # Get all users
    users = CustomUser.objects.all().select_related('department')
    
    # Get all systems
    systems = System.objects.filter(is_active=True).order_by('name')
    
    # Apply filters
    if user_id:
        users = users.filter(id=user_id)
    if department_id:
        users = users.filter(department_id=department_id)
    if search:
        users = users.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Get all access assignments with system usernames
    access_assignments = UserSystemAccess.objects.filter(
        status__in=['Active', 'Approved']
    ).select_related('user', 'user__department', 'system', 'username_verified_by')
    
    if system_id:
        access_assignments = access_assignments.filter(system_id=system_id)
    
    export_format = request.GET.get('export')
    if export_format in {'csv', 'xlsx'}:
        ordered_assignments = access_assignments.order_by(
            'user__first_name', 'user__last_name', 'system__name'
        )
        rows = build_cross_system_mapping_rows(ordered_assignments, request)
        if export_format == 'xlsx':
            return export_cross_system_mapping_to_excel(rows)
        return export_cross_system_mapping_to_csv(rows)

    # Build mapping: user_id -> {system_id: details}
    user_system_mapping = {}
    for access in access_assignments:
        user_id = access.user_id
        system_id = access.system_id
        # Use effective_username property which handles system_username and access_username fallback
        username = access.effective_username
        artifact_file_url = ''
        if access.username_verification_artifact:
            try:
                artifact_file_url = access.username_verification_artifact.url
            except ValueError:
                artifact_file_url = ''
        
        if user_id not in user_system_mapping:
            user_system_mapping[user_id] = {}
        user_system_mapping[user_id][system_id] = {
            'username': username,
            'access_type': access.access_type,
            'status': access.status,
            'is_generic': access.is_generic_account,
            'access_id': access.id,
            'verified_by': access.username_verified_by.full_name if access.username_verified_by else '',
            'verified_date': access.username_verified_date,
            'artifact_file_url': artifact_file_url,
            'artifact_external_url': access.username_verification_artifact_url or '',
            'has_verification': access.has_username_verification,
            'system_owner_approved': access.system_owner_approved,
            'system_owner_approval_date': access.system_owner_approval_date,
            'system_owner_name': access.system.system_owner.full_name if access.system and access.system.system_owner else '',
            'legitimate_business_need': access.legitimate_business_need or '',
            'business_justification': access.business_justification or '',
        }
    
    # Filter users if show_only_with_access
    if show_only_with_access:
        users = users.filter(id__in=list(user_system_mapping.keys()))
    
    # Order users
    users = users.order_by('first_name', 'last_name')
    
    # Pagination
    paginator = Paginator(users, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_users = users.filter(exclude_from_metrics=False).count()
    users_with_access = len(user_system_mapping)
    total_systems = systems.count()
    
    context = {
        'page_obj': page_obj,
        'users': page_obj,
        'systems': systems,
        'user_system_mapping': user_system_mapping,
        'selected_user_id': user_id,
        'selected_system_id': system_id,
        'selected_department_id': department_id,
        'search': search,
        'show_only_with_access': show_only_with_access,
        'total_users': total_users,
        'users_with_access': users_with_access,
        'total_systems': total_systems,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
    }
    
    return render(request, 'access_management/cross_system_account_mapping.html', context)


@login_required
def user_cross_system_accounts(request, user_id):
    """Show all usernames for a single employee across all systems"""
    user = get_object_or_404(CustomUser, pk=user_id)
    
    # Get all access assignments for this user
    access_assignments = UserSystemAccess.objects.filter(
        user=user
    ).select_related('system').order_by('system__name')
    
    # Build mapping: system -> access details
    system_accounts = []
    for access in access_assignments:
        artifact_file_url = ''
        if access.username_verification_artifact:
            try:
                artifact_file_url = access.username_verification_artifact.url
            except ValueError:
                artifact_file_url = ''

        system_accounts.append({
            'system': access.system,
            'system_username': access.effective_username or 'N/A',
            'access_type': access.access_type,
            'status': access.status,
            'is_generic': access.is_generic_account,
            'generic_remediated': access.generic_account_remediated,
            'access_start_date': access.access_start_date,
            'access_end_date': access.access_end_date,
            'access_id': access.id,
            'verified_by': access.username_verified_by.full_name if access.username_verified_by else '',
            'verified_date': access.username_verified_date,
            'artifact_file_url': artifact_file_url,
            'artifact_external_url': access.username_verification_artifact_url or '',
            'has_verification': access.has_username_verification,
        })
    
    # Get all systems to show which ones user doesn't have access to
    all_systems = System.objects.filter(is_active=True).order_by('name')
    systems_with_access = {acc['system'].id for acc in system_accounts}
    systems_without_access = [sys for sys in all_systems if sys.id not in systems_with_access]
    
    context = {
        'user': user,
        'system_accounts': system_accounts,
        'systems_without_access': systems_without_access,
        'total_systems': all_systems.count(),
        'systems_with_access_count': len(system_accounts),
    }
    
    return render(request, 'access_management/user_cross_system_accounts.html', context)


@login_required
def accounts_status(request):
    """Show service accounts split by assignment status (assigned vs unassigned).

    This view prepares simple dicts to match the `accounts_status.html` template
    so the template does not depend on a specific model API.
    """
    from service_accounts.models import ServiceAccount
    assigned_qs = ServiceAccount.objects.filter(
        Q(owner__isnull=False) | Q(admin_user__isnull=False)
    ).select_related('system', 'owner', 'admin_user').order_by('system__name', 'account_name')
    unassigned_qs = ServiceAccount.objects.filter(
        owner__isnull=True, admin_user__isnull=True
    ).select_related('system').order_by('system__name', 'account_name')

    assigned_accounts = []
    for s in assigned_qs:
        assigned_accounts.append({
            'display_name': s.account_name,
            'username': s.account_name,
            'system': s.system,
            'assigned_to': s.owner.get_full_name() if s.owner else (s.admin_user.get_full_name() if s.admin_user else None),
            'id': s.id,
        })

    unassigned_accounts = []
    for s in unassigned_qs:
        unassigned_accounts.append({
            'display_name': s.account_name,
            'username': s.account_name,
            'system': s.system,
            'assigned_to': None,
            'id': s.id,
        })

    context = {
        'assigned_accounts': assigned_accounts,
        'unassigned_accounts': unassigned_accounts,
    }

    return render(request, 'access_management/accounts_status.html', context)
