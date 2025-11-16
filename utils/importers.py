import csv
import io
import re
from typing import Iterable, Tuple, List, Dict, Any, Iterator

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils.dateparse import parse_date

from openpyxl import load_workbook

from departments.models import Department
from systems.models import System
from hardware.models import HardwareAsset

User = get_user_model()


class ImportErrorCollection(Exception):
    """Aggregate multiple validation errors for CSV imports."""

    def __init__(self, errors: Iterable[str]):
        self.errors = list(errors)
        message = '\n'.join(self.errors)
        super().__init__(message)


def _csv_reader(uploaded_file) -> csv.DictReader:
    """
    Normalise uploaded files/in-memory file-like objects into a CSV DictReader.
    Ensures UTF-8 decoding (with BOM support) and positions the pointer at the start.
    Attempts to auto-detect common delimiters used by Active Directory exports.
    """
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
    raw_data = uploaded_file.read()

    if isinstance(raw_data, bytes):
        raw_data = raw_data.decode('utf-8-sig', errors='replace')
    elif not isinstance(raw_data, str):
        raw_data = str(raw_data)

    if raw_data.startswith('#TYPE'):
        raw_data = '\n'.join(
            line for line in raw_data.splitlines() if not line.startswith('#TYPE')
        )

    sample = raw_data[:4096]
    delimiter = ','
    dialect = None
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t', '|'])
    except csv.Error:
        if sample.count('\t') > sample.count(delimiter):
            delimiter = '\t'
        elif sample.count(';') > sample.count(delimiter):
            delimiter = ';'
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    buffer = io.StringIO(raw_data)
    if dialect:
        return csv.DictReader(buffer, dialect=dialect)
    return csv.DictReader(buffer, delimiter=delimiter)


def _excel_dict_rows(uploaded_file) -> List[Tuple[int, Dict[str, Any]]]:
    """
    Convert an uploaded Excel file into a list of (row_number, row_dict) tuples.
    """
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
    raw_data = uploaded_file.read()

    if isinstance(raw_data, str):
        raw_data = raw_data.encode('utf-8')

    workbook = load_workbook(filename=io.BytesIO(raw_data), data_only=True, read_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []

    headers = [(_clean_value(header) or '').strip() for header in headers]
    normalized_rows: List[Tuple[int, Dict[str, Any]]] = []
    for excel_index, row_values in enumerate(rows_iter, start=2):
        row_dict: Dict[str, Any] = {}
        is_empty = True
        for header, value in zip(headers, row_values):
            if not header:
                continue
            cleaned_value = _clean_value(value)
            if cleaned_value not in ('', None):
                is_empty = False
            row_dict[header] = cleaned_value
        if is_empty:
            continue
        normalized_rows.append((excel_index, row_dict))

    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    return normalized_rows


def _clean_value(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _normalize_header_key(header: str) -> str:
    header = (header or '').strip().lower()
    header = re.sub(r'\(.*?\)', '', header)
    header = re.sub(r'[^a-z0-9]+', '_', header)
    return header.strip('_')


USER_FIELD_ALIASES = {
    'user_logon_name': 'username',
    'user_logon_name_pre_windows_2000': 'username',
    'samaccountname': 'username',
    'user_id': 'username',
    'user_principal_name': 'email',
    'userprincipalname': 'email',
    'mail': 'email',
    'email_address': 'email',
    'work_email': 'email',
    'givenname': 'first_name',
    'first_name': 'first_name',
    'sn': 'last_name',
    'surname': 'last_name',
    'last_name': 'last_name',
    'telephonenumber': 'phone_primary',
    'telephone_number': 'phone_primary',
    'business_phone': 'phone_primary',
    'work_phone': 'phone_primary',
    'office_phone': 'phone_primary',
    'mobile': 'phone_secondary',
    'mobile_phone': 'phone_secondary',
    'home_phone': 'phone_secondary',
    'pager': 'phone_secondary',
    'title': 'position',
    'job_title': 'position',
    'department': 'department_name',
    'employeeid': 'employee_id',
    'employee_number': 'employee_id',
    'employeenumber': 'employee_id',
    'employee_no': 'employee_id',
    'account_active': 'is_active',
    'enabled': 'is_active',
    'account_enabled': 'is_active',
    'physical_delivery_office_name': 'office_location',
    'office': 'office_location',
    'street_address': 'work_address',
    'address': 'work_address',
    'state': 'state_province',
    'state_or_province': 'state_province',
    'zip': 'postal_code',
    'zip_code': 'postal_code',
    'country_region': 'country',
    'co': 'country',
    'description': 'description',
    'comment': 'notes',
    # Reports to aliases
    'reports_to_username': 'reports_to__username',
    'reports_to_employee_id': 'reports_to__employee_id',
    'reports_to_first_name': 'reports_to__first_name',
    'reports_to_last_name': 'reports_to__last_name',
    'manager_username': 'reports_to__username',
    'manager_employee_id': 'reports_to__employee_id',
    'manager': 'reports_to__username',
    # Department aliases (already handled separately but adding for completeness)
    'dept_code': 'department_code',
    'dept_name': 'department_name',
    # Employee level aliases
    'level': 'employee_level',
    'employee_level': 'employee_level',
}


def _normalise_user_row(row: Dict[str, Any]) -> Dict[str, Any]:
    normalised: Dict[str, Any] = {}
    interim: Dict[str, Any] = {}

    for key, value in row.items():
        normalised_key = _normalize_header_key(key)
        if not normalised_key:
            continue
        if normalised_key not in interim or interim[normalised_key] in ('', None):
            interim[normalised_key] = _clean_value(value)

    for key, value in interim.items():
        target_key = USER_FIELD_ALIASES.get(key, key)
        if target_key not in normalised or normalised[target_key] in ('', None):
            normalised[target_key] = value

    return normalised


def _populate_name_from_display_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Populate missing first_name/last_name values using display-style columns
    commonly present in Active Directory exports.
    """
    first = (row.get('first_name') or '').strip()
    last = (row.get('last_name') or '').strip()
    if first and last:
        return row

    display_candidates = (
        row.get('display_name'),
        row.get('name'),
        row.get('full_name'),
        row.get('fullname'),
    )
    display_value = next((value for value in display_candidates if value), None)
    if not display_value:
        return row

    display_value = str(display_value).strip()
    if not display_value:
        return row

    extracted_first = ''
    extracted_last = ''

    if ',' in display_value:
        last_part, remainder = [part.strip() for part in display_value.split(',', 1)]
        extracted_last = last_part or extracted_last
        remainder_parts = [part for part in re.split(r'\s+', remainder) if part]
        if remainder_parts:
            extracted_first = remainder_parts[0]
            if len(remainder_parts) > 1 and not extracted_last:
                extracted_last = remainder_parts[-1]
    else:
        parts = [part for part in re.split(r'\s+', display_value) if part]
        if len(parts) == 1:
            extracted_first = extracted_first or parts[0]
        elif len(parts) >= 2:
            extracted_first = parts[0]
            extracted_last = parts[-1]

    if not first and extracted_first:
        row['first_name'] = extracted_first
    if not last and extracted_last:
        row['last_name'] = extracted_last

    return row


def _sanitize_username(candidate: str) -> str:
    candidate = (candidate or '').strip().lower()
    candidate = re.sub(r'[^a-z0-9._-]+', '', candidate)
    return candidate


def _derive_username(row: Dict[str, Any]) -> str:
    username = _sanitize_username(row.get('username', ''))
    if username:
        return username

    email = (row.get('email') or '').strip()
    if email and '@' in email:
        user_part = email.split('@', 1)[0]
        username = _sanitize_username(user_part)
        if username:
            return username

    employee_id = _sanitize_username(row.get('employee_id', ''))
    if employee_id:
        return employee_id

    first = _sanitize_username(row.get('first_name', ''))
    last = _sanitize_username(row.get('last_name', ''))
    if first and last:
        username = _sanitize_username(f"{first}.{last}")
        if username:
            return username
    if first:
        return first
    if last:
        return last

    display = (row.get('display_name') or row.get('name') or row.get('full_name') or row.get('fullname') or '').strip()
    if display:
        display = re.sub(r'\s+', ' ', display)
        candidate = display.split(' ')[0]
        username = _sanitize_username(candidate)
        if username:
            return username

    return 'user'


def _iter_user_rows(uploaded_file) -> Iterator[Tuple[int, Dict[str, Any]]]:
    name = getattr(uploaded_file, 'name', '') or ''
    lower_name = name.lower()

    excel_extensions = ('.xlsx', '.xlsm', '.xls', '.xlsb')
    if lower_name.endswith(excel_extensions):
        rows = _excel_dict_rows(uploaded_file)
        for index, row in rows:
            yield index, _populate_name_from_display_fields(_normalise_user_row(row))
        return

    # Fallback to signature detection for Excel files without an extension.
    excel_signature = None
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
        excel_signature = uploaded_file.read(4)
        uploaded_file.seek(0)
    if isinstance(excel_signature, bytes) and excel_signature.startswith(b'PK\x03\x04'):
        rows = _excel_dict_rows(uploaded_file)
        for index, row in rows:
            yield index, _populate_name_from_display_fields(_normalise_user_row(row))
        return

    reader = _csv_reader(uploaded_file)
    for index, row in enumerate(reader, start=2):
        yield index, _populate_name_from_display_fields(_normalise_user_row(row))


def _parse_bool(value, default=False):
    if value in (None, ''):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'true', '1', 'yes', 'y', 't'}


def _parse_int(value, default=None):
    if value in (None, ''):
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def _collect_row_errors(row_number: int, errors: Iterable[str]) -> str:
    return f"Row {row_number}: " + '; '.join(errors)


def _generate_unique_employee_id() -> str:
    """
    Generate a unique employee ID with the prefix 'EMP'.
    Ensures uniqueness against existing users.
    """
    prefix = 'EMP'
    counter = User.objects.filter(employee_id__startswith=prefix).count() + 1
    candidate = f"{prefix}{counter:06d}"
    while User.objects.filter(employee_id=candidate).exists():
        counter += 1
        candidate = f"{prefix}{counter:06d}"
    return candidate

def _generate_unique_username(base_username: str, keep_employee_id: str) -> str:
    """
    Ensure username is unique. If the base exists for a different employee,
    append an incrementing numeric suffix until it's unique.
    """
    base = (base_username or '').strip()
    if not base:
        base = 'user'
    # If username belongs to the same employee, allow it
    existing = User.objects.filter(username=base).first()
    if existing and getattr(existing, 'employee_id', None) == keep_employee_id:
        return base
    if not User.objects.filter(username=base).exclude(employee_id=keep_employee_id).exists():
        return base
    counter = 1
    while True:
        candidate = f"{base}{counter}"
        if not User.objects.filter(username=candidate).exclude(employee_id=keep_employee_id).exists():
            return candidate
        counter += 1

def import_users_from_csv(file):
    """Import or update user records from a CSV or Excel file."""
    # Relax required columns; only 'username' is mandatory (aliases allowed).
    required_columns = {'username'}
    errors = []

    with transaction.atomic():
        for index, row in _iter_user_rows(file):
            if not row.get('username'):
                derived_username = _derive_username(row)
                if derived_username:
                    row['username'] = derived_username

            missing_columns = [col for col in required_columns if not row.get(col)]
            if missing_columns:
                errors.append(_collect_row_errors(index, [f"Missing required columns: {', '.join(missing_columns)}"]))
                continue

            department = None
            department_code = row.get('department_code')
            department_name = row.get('department_name')
            if department_code:
                department = Department.objects.filter(code=department_code).first()
                if department is None:
                    errors.append(_collect_row_errors(index, [f"Department with code '{department_code}' not found"]))
                    continue
            elif department_name:
                department = Department.objects.filter(name__iexact=department_name).first()
                if department is None:
                    errors.append(_collect_row_errors(index, [f"Department with name '{department_name}' not found"]))
                    continue

            # Handle reports_to relationship
            reports_to = None
            reports_to_username = row.get('reports_to__username') or row.get('reports_to_username')
            reports_to_employee_id = row.get('reports_to__employee_id') or row.get('reports_to_employee_id')
            reports_to_first_name = row.get('reports_to__first_name') or row.get('reports_to_first_name')
            reports_to_last_name = row.get('reports_to__last_name') or row.get('reports_to_last_name')
            
            if reports_to_username:
                reports_to = User.objects.filter(username=reports_to_username).first()
                if reports_to is None:
                    # Warning only, don't stop import
                    pass
            elif reports_to_employee_id:
                reports_to = User.objects.filter(employee_id=reports_to_employee_id).first()
                if reports_to is None:
                    # Warning only, don't stop import
                    pass
            elif reports_to_first_name and reports_to_last_name:
                reports_to = User.objects.filter(
                    first_name__iexact=reports_to_first_name.strip(),
                    last_name__iexact=reports_to_last_name.strip()
                ).first()
                if reports_to is None:
                    # Warning only, don't stop import
                    pass

            join_date = parse_date(row.get('join_date') or '')
            end_date = parse_date(row.get('end_date') or '')
            # Note: date_joined is auto-generated by Django, cannot be set via import

            # Defaults for missing values
            username = row['username'].strip()
            first_name = (row.get('first_name') or 'User').strip()
            last_name = (row.get('last_name') or 'Unknown').strip()
            email = (row.get('email') or f"{username}@example.com").strip()
            phone_primary = (row.get('phone_primary') or '+10000000000').strip()
            position = (row.get('position') or 'Staff').strip()
            employee_id = (row.get('employee_id') or '').strip() or _generate_unique_employee_id()
            # Ensure username is unique across different employees
            username = _generate_unique_username(username, employee_id)

            defaults = {
                'username': username,
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'employee_id': employee_id,
                'phone_primary': phone_primary,
                'position': position,
                'employment_type': (row.get('employment_type') or 'Full-time').strip(),
                'employment_status': (row.get('employment_status') or 'Active').strip(),
                'employee_level': (row.get('employee_level') or 'Entry').strip(),
                'department': department,
                'reports_to': reports_to,
                'is_active': _parse_bool(row.get('is_active'), default=True),
            }

            # All optional fields that can be imported (matching custom export fields)
            optional_fields = [
                'phone_secondary',
                'personal_email',
                'job_title',
                'description',
                'office_location',
                'office_room',
                'work_address',
                'city',
                'state_province',
                'country',
                'postal_code',
                'notes',
            ]

            for field_name in optional_fields:
                if row.get(field_name) not in (None, ''):
                    defaults[field_name] = row[field_name].strip()

            if join_date:
                defaults['join_date'] = join_date
            if end_date:
                defaults['end_date'] = end_date

            user, created = User.objects.update_or_create(
                employee_id=defaults['employee_id'],
                defaults=defaults,
            )

            password = row.get('password')
            if password:
                user.set_password(password)
                user.save(update_fields=['password'])
            elif created and not user.has_usable_password():
                user.set_unusable_password()
                user.save(update_fields=['password'])

        if errors:
            raise ImportErrorCollection(errors)


def import_departments_from_csv(file):
    """Import or update department records from a CSV file."""
    reader = _csv_reader(file)
    # department_type is optional; defaults to 'Department' if omitted.
    required_columns = {'name', 'code'}
    errors = []
    # Store (child_department, parent_code, parent_name) for post-processing
    parent_links: List[Tuple[Department, str, str]] = []

    with transaction.atomic():
        for index, row in enumerate(reader, start=2):
            missing_columns = [col for col in required_columns if not row.get(col)]
            if missing_columns:
                errors.append(_collect_row_errors(index, [f"Missing required columns: {', '.join(missing_columns)}"]))
                continue

            code = (row['code'] or '').strip()
            name = (row['name'] or '').strip()
            raw_department_type = (row.get('department_type') or '').strip()
            # Validate department_type if provided
            if raw_department_type and raw_department_type not in dict(Department.DEPARTMENT_TYPE_CHOICES):
                allowed = ', '.join(dict(Department.DEPARTMENT_TYPE_CHOICES).keys())
                errors.append(_collect_row_errors(index, [f"Invalid department_type '{raw_department_type}'. Allowed: {allowed}"]))
                continue

            # Pre-check for unique name conflicts (model enforces unique=True on name)
            # If another department already uses this name with a different code, report a readable error.
            existing_with_name = Department.objects.filter(name=name).exclude(code=code).first()
            if existing_with_name:
                errors.append(_collect_row_errors(
                    index,
                    [f"Department name '{name}' already exists with code '{existing_with_name.code}'. Use a unique name or match the existing code."]
                ))
                continue

            defaults = {
                'name': name,
                'description': (row.get('description') or '').strip(),
                'department_type': raw_department_type or 'Department',
                'is_active': _parse_bool(row.get('is_active'), default=True),
            }

            optional_fields = [
                'cost_center',
                'budget_code',
                'office_location',
                'phone',
                'email',
            ]
            for field_name in optional_fields:
                if row.get(field_name) not in (None, ''):
                    defaults[field_name] = row[field_name].strip()

            department, _ = Department.objects.update_or_create(
                code=code,
                defaults=defaults,
            )

            parent_code = (row.get('parent_department_code') or '').strip()
            parent_name = (row.get('parent_department_name') or '').strip()
            if parent_code or parent_name:
                parent_links.append((department, parent_code, parent_name))

        # Resolve parent relationships outside the main loop to ensure all departments exist.
        if not errors and parent_links:
            for department, parent_code, parent_name in parent_links:
                parent = None
                # 1) Try by code if provided
                if parent_code:
                    parent = Department.objects.filter(code=parent_code).first()
                # 2) Try by name if not found and name provided
                if parent is None and parent_name:
                    parent = Department.objects.filter(name=parent_name).first()
                # 3) If still not found, try interpreting provided code as a name (common CSV confusion)
                if parent is None and parent_code:
                    parent = Department.objects.filter(name=parent_code).first()
                # 4) If still not found but we have a name, auto-create parent
                if parent is None and parent_name:
                    # Choose code: prefer provided parent_code, else use parent_name as code if unique, otherwise skip to slug-like variant
                    candidate_code = parent_code or parent_name
                    candidate_code = candidate_code.strip()
                    # Ensure uniqueness of code
                    if Department.objects.filter(code=candidate_code).exists():
                        # Append incremental suffix to avoid collision
                        base = candidate_code[:45]  # leave room for suffix
                        suffix_idx = 1
                        new_code = f"{base}-{suffix_idx}"
                        while Department.objects.filter(code=new_code).exists():
                            suffix_idx += 1
                            new_code = f"{base}-{suffix_idx}"
                        candidate_code = new_code
                    parent = Department.objects.create(
                        name=parent_name,
                        code=candidate_code,
                        department_type='Department',
                        is_active=True,
                    )
                # 5) If still not found and only a parent_code was given (likely a human-readable name used as code), auto-create using it.
                if parent is None and parent_code and not parent_name:
                    candidate_code = parent_code.strip()
                    candidate_name = parent_code.strip()
                    if Department.objects.filter(code=candidate_code).exists():
                        base = candidate_code[:45]
                        suffix_idx = 1
                        new_code = f"{base}-{suffix_idx}"
                        while Department.objects.filter(code=new_code).exists():
                            suffix_idx += 1
                            new_code = f"{base}-{suffix_idx}"
                        candidate_code = new_code
                    parent = Department.objects.create(
                        name=candidate_name,
                        code=candidate_code,
                        department_type='Department',
                        is_active=True,
                    )
                if parent is None:
                    ref = parent_code or parent_name or '(unspecified)'
                    errors.append(f"Parent department with reference '{ref}' not found for department '{department.code}'")
                    continue
                if department.parent_department_id != parent.id:
                    department.parent_department = parent
                    department.save(update_fields=['parent_department'])

        if errors:
            raise ImportErrorCollection(errors)


def import_systems_from_csv(file):
    """Import or update system records from a CSV file."""
    reader = _csv_reader(file)
    required_columns = {'name', 'code', 'system_type', 'criticality_level', 'environment_type'}
    errors = []

    with transaction.atomic():
        for index, row in enumerate(reader, start=2):
            missing_columns = [col for col in required_columns if not row.get(col)]
            if missing_columns:
                errors.append(_collect_row_errors(index, [f"Missing required columns: {', '.join(missing_columns)}"]))
                continue

            system_owner = None
            owner_employee_id = row.get('system_owner_employee_id')
            if owner_employee_id:
                system_owner = User.objects.filter(employee_id=owner_employee_id.strip()).first()
                if system_owner is None:
                    errors.append(_collect_row_errors(index, [f"System owner with employee ID '{owner_employee_id}' not found"]))
                    continue

            technical_lead = None
            technical_lead_employee_id = row.get('technical_lead_employee_id')
            if technical_lead_employee_id:
                technical_lead = User.objects.filter(employee_id=technical_lead_employee_id.strip()).first()
                if technical_lead is None:
                    errors.append(_collect_row_errors(index, [f"Technical lead with employee ID '{technical_lead_employee_id}' not found"]))
                    continue

            defaults = {
                'name': row['name'].strip(),
                'description': (row.get('description') or '').strip(),
                'system_type': row['system_type'].strip(),
                'criticality_level': row['criticality_level'].strip(),
                'environment_type': row['environment_type'].strip(),
                'system_owner': system_owner,
                'technical_lead': technical_lead,
                'requires_approval': _parse_bool(row.get('requires_approval'), default=True),
                'is_active': _parse_bool(row.get('is_active'), default=True),
                'is_monitored': _parse_bool(row.get('is_monitored'), default=True),
            }

            optional_fields = [
                'url',
                'ip_address',
                'server_name',
                'version',
                'vendor',
                'vendor_contact',
                'support_contact',
                'documentation_url',
                'authentication_type',
                'data_classification',
                'backup_frequency',
                'maintenance_window',
            ]

            for field_name in optional_fields:
                if row.get(field_name) not in (None, ''):
                    defaults[field_name] = row[field_name].strip()

            System.objects.update_or_create(
                code=row['code'].strip(),
                defaults=defaults,
            )

        if errors:
            raise ImportErrorCollection(errors)


def import_hardware_from_csv(file):
    """Import or update hardware asset records from a CSV file."""
    reader = _csv_reader(file)
    required_columns = {'name', 'asset_tag', 'hardware_type', 'status'}
    errors = []

    valid_hardware_types = dict(HardwareAsset.HARDWARE_TYPE_CHOICES).keys()
    valid_statuses = dict(HardwareAsset.STATUS_CHOICES).keys()

    with transaction.atomic():
        for index, row in enumerate(reader, start=2):
            missing_columns = [col for col in required_columns if not row.get(col)]
            if missing_columns:
                errors.append(_collect_row_errors(index, [f"Missing required columns: {', '.join(missing_columns)}"]))
                continue

            hardware_type = (row['hardware_type'] or '').strip()
            if hardware_type not in valid_hardware_types:
                errors.append(_collect_row_errors(index, [f"Invalid hardware_type '{hardware_type}'."]))
                continue

            status = (row['status'] or '').strip()
            if status not in valid_statuses:
                errors.append(_collect_row_errors(index, [f"Invalid status '{status}'."]))
                continue

            department = None
            department_code = (row.get('department_code') or '').strip()
            if department_code:
                department = Department.objects.filter(code=department_code).first()
                if department is None:
                    errors.append(_collect_row_errors(index, [f"Department with code '{department_code}' not found"]))
                    continue

            primary_user = None
            primary_user_employee_id = (row.get('primary_user_employee_id') or '').strip()
            if primary_user_employee_id:
                primary_user = User.objects.filter(employee_id=primary_user_employee_id).first()
                if primary_user is None:
                    errors.append(_collect_row_errors(index, [f"Primary user with employee ID '{primary_user_employee_id}' not found"]))
                    continue

            purchase_date = parse_date(row.get('purchase_date') or '')
            warranty_expiration = parse_date(row.get('warranty_expiration') or '')
            end_of_life_date = parse_date(row.get('end_of_life_date') or '')
            last_inventory_check = parse_date(row.get('last_inventory_check') or '')
            next_inventory_check = parse_date(row.get('next_inventory_check') or '')

            defaults = {
                'name': (row['name'] or '').strip(),
                'hardware_type': hardware_type,
                'status': status,
                'serial_number': (row.get('serial_number') or '').strip() or None,
                'manufacturer': (row.get('manufacturer') or '').strip() or None,
                'model_number': (row.get('model_number') or '').strip() or None,
                'operating_system': (row.get('operating_system') or '').strip() or None,
                'cpu': (row.get('cpu') or '').strip() or None,
                'memory_gb': _parse_int(row.get('memory_gb')),
                'storage_capacity_gb': _parse_int(row.get('storage_capacity_gb')),
                'location': (row.get('location') or '').strip() or None,
                'ip_address': (row.get('ip_address') or '').strip() or None,
                'mac_address': (row.get('mac_address') or '').strip() or None,
                'notes': (row.get('notes') or '').strip() or None,
                'is_virtual': _parse_bool(row.get('is_virtual'), default=False),
                'requires_patch_management': _parse_bool(row.get('requires_patch_management'), default=True),
                'department': department,
                'primary_user': primary_user,
            }

            if purchase_date:
                defaults['purchase_date'] = purchase_date
            if warranty_expiration:
                defaults['warranty_expiration'] = warranty_expiration
            if end_of_life_date:
                defaults['end_of_life_date'] = end_of_life_date
            if last_inventory_check:
                defaults['last_inventory_check'] = last_inventory_check
            if next_inventory_check:
                defaults['next_inventory_check'] = next_inventory_check

            asset, _ = HardwareAsset.objects.update_or_create(
                asset_tag=(row['asset_tag'] or '').strip(),
                defaults=defaults,
            )

            # Assigned users (semicolon-separated employee IDs)
            assigned_users_raw = row.get('assigned_user_employee_ids') or ''
            assigned_user_ids = []
            if assigned_users_raw:
                missing_users = []
                for employee_id in assigned_users_raw.split(';'):
                    employee_id = employee_id.strip()
                    if not employee_id:
                        continue
                    user = User.objects.filter(employee_id=employee_id).first()
                    if user is None:
                        missing_users.append(employee_id)
                    else:
                        assigned_user_ids.append(user.id)
                if missing_users:
                    errors.append(_collect_row_errors(index, [f"Assigned user(s) with employee IDs {', '.join(missing_users)} not found"]))
                    continue
            if assigned_user_ids:
                asset.assigned_users.set(assigned_user_ids)
            else:
                asset.assigned_users.clear()

            # Related systems (semicolon-separated system codes)
            related_systems_raw = row.get('related_system_codes') or ''
            related_system_ids = []
            if related_systems_raw:
                missing_systems = []
                for system_code in related_systems_raw.split(';'):
                    system_code = system_code.strip()
                    if not system_code:
                        continue
                    system = System.objects.filter(code=system_code).first()
                    if system is None:
                        missing_systems.append(system_code)
                    else:
                        related_system_ids.append(system.id)
                if missing_systems:
                    errors.append(_collect_row_errors(index, [f"Related system(s) with code(s) {', '.join(missing_systems)} not found"]))
                    continue
            if related_system_ids:
                asset.related_systems.set(related_system_ids)
            else:
                asset.related_systems.clear()

        if errors:
            raise ImportErrorCollection(errors)